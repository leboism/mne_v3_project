"""Export Excel fichier de notes depuis l'onglet Résultats."""

from __future__ import annotations

from pathlib import Path
import tempfile

from openpyxl import load_workbook

from mne_grade_manager.core.database import Database
from mne_grade_manager.services.repository import Repository
from mne_grade_manager.services.results_excel import (
    grade_excel_value,
    suggest_notes_workbook_filename,
    write_results_notes_workbook,
)


def _repo() -> Repository:
    db = Database(Path(tempfile.mkdtemp()) / "results_excel.sqlite3")
    return Repository(db)


def _setup_minimal(repo: Repository) -> None:
    cid = repo.add_course("D4NM12A", "Thermodynamics", ects=3)
    repo.add_assessment(cid, "EE", "EE", 0.4, session=1)
    repo.add_assessment(cid, "EEF", "EEF", 0.6, session=1)
    repo.add_assessment(cid, "EE", "EE", 1.0, session=2)
    tid_p = repo.add_template("M1 P", "M1", "P", "2025-2026", "1")
    tid_c = repo.add_template("M1 C", "M1", "C", "2025-2026", "1")
    repo.add_course_to_template(tid_p, cid, block_name="Block 1")
    repo.add_course_to_template(tid_c, cid, block_name="Block 1")
    sid_p = repo.add_student(
        "INE-P1", "", "", "Dupont", "Alice", academic_year="2025-2026", level="M1", track="P"
    )
    sid_c = repo.add_student(
        "INE-C1", "", "", "Martin", "Bob", academic_year="2025-2026", level="M1", track="C"
    )
    repo.enroll_student(sid_p, tid_p)
    repo.enroll_student(sid_c, tid_c)
    aid = int(repo.get_grades_for_student_course(sid_p, cid)[0]["assessment_id"])
    repo.upsert_grade(sid_p, aid, 12.5, status="OK")
    repo.upsert_grade(sid_c, aid, 14.0, status="OK")


def test_suggest_notes_workbook_filename() -> None:
    assert suggest_notes_workbook_filename(academic_year="2025-2026", level="M1") == (
        "Fichier_de_notes_M1NE_2025-26.xlsx"
    )


def test_grade_excel_value_status_and_numeric() -> None:
    assert grade_excel_value(None, "ABJ") == "ABJ"
    assert grade_excel_value(12.5, "OK") == 12.5
    assert grade_excel_value(None, "DEF", assessment_session=2, assessment_kind="EE") == "DEF"
    assert grade_excel_value(
        None, "DEF", assessment_session=2, assessment_kind="CC", assessment_name="CC Rep (30%)"
    ) is None


def test_write_results_notes_workbook_structure() -> None:
    repo = _repo()
    _setup_minimal(repo)
    path = Path(tempfile.mkdtemp()) / "notes.xlsx"
    write_results_notes_workbook(
        repo,
        academic_year="2025-2026",
        level="M1",
        path=path,
    )
    wb = load_workbook(path, data_only=True)
    assert "Liste Etudiants" in wb.sheetnames
    assert "session 1" in wb.sheetnames
    assert "session 2" in wb.sheetnames
    course_sheets = [
        s for s in wb.sheetnames if s not in {"Liste Etudiants", "session 1", "session 2"}
    ]
    assert course_sheets
    ws_list = wb["Liste Etudiants"]
    assert ws_list.cell(3, 1).value == "Numéro"
    assert ws_list.cell(4, 2).value == "Dupont"
    ws_course = wb[course_sheets[0]]
    assert "Thermodynamics" in str(ws_course.cell(1, 5).value)
    assert ws_course.cell(3, 5).value == "EE (0,4)"
    assert ws_course.cell(3, 7).value == "Note S1"
    dupont_row = next(
        r
        for r in range(4, ws_course.max_row + 1)
        if ws_course.cell(r, 2).value == "Dupont"
    )
    assert ws_course.cell(dupont_row, 5).value == 12.5
    assert ws_course.cell(dupont_row, 9).value == -1


def test_m1_export_excludes_students_already_in_m2() -> None:
    repo = _repo()
    cid = repo.add_course("D4NM12A", "Thermodynamics", ects=3)
    repo.add_assessment(cid, "EE", "EE", 1.0, session=1)
    tid = repo.add_template("M1 P", "M1", "P", "2025-2026", "1")
    repo.add_course_to_template(tid, cid, block_name="Block 1")
    sid_m1 = repo.add_student(
        "INE-M1", "", "", "Reste", "Marie", academic_year="2025-2026", level="M1", track="P"
    )
    sid_m2 = repo.add_student(
        "INE-M2", "", "", "Parti", "Paul", academic_year="2026-2027", level="M2", track="P"
    )
    repo.enroll_student(sid_m1, tid)
    repo.enroll_student(sid_m2, tid)
    path = Path(tempfile.mkdtemp()) / "notes.xlsx"
    write_results_notes_workbook(
        repo, academic_year="2025-2026", level="M1", path=path
    )
    wb = load_workbook(path, data_only=True)
    names = {
        (ws.cell(r, 2).value, ws.cell(r, 3).value)
        for ws in (wb["Liste Etudiants"],)
        for r in range(4, ws.max_row + 1)
        if ws.cell(r, 2).value
    }
    assert ("Reste", "Marie") in names
    assert ("Parti", "Paul") not in names
