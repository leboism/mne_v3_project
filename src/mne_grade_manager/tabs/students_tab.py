from __future__ import annotations

from datetime import date
from pathlib import Path
from typing import Any

from PySide6.QtCore import Qt
from PySide6.QtGui import QBrush, QColor
from PySide6.QtWidgets import (
    QAbstractItemView,
    QButtonGroup,
    QComboBox,
    QDialog,
    QDialogButtonBox,
    QFileDialog,
    QGridLayout,
    QHBoxLayout,
    QLabel,
    QLineEdit,
    QMenu,
    QMessageBox,
    QPushButton,
    QRadioButton,
    QTableWidget,
    QTableWidgetItem,
    QToolButton,
    QVBoxLayout,
    QWidget,
)

from ..core.master_team import MNE_ENROLLMENT_INSTITUTIONS
from ..core.parcours import PARCOURS_BY_LEVEL, track_display_label
from ..services.student_mobility import is_erasmus_student
from ..gui.admission_import_dialog import AdmissionImportDialog, import_admission_dossiers
from ..gui.dialogs import StudentDialog
from ..gui.progression_dialog import StudentProgressionDialog
from ..gui.student_email_list_dialog import StudentEmailListDialog
from ..gui.student_profile_dialog import StudentProfileDialog
from ..services.admission_import import collect_admission_pdfs, parse_admission_pdf
from ..services.dates import format_age_display, normalize_birth_date_iso
from ..services.student_excel import (
    STUDENT_EXPORT_FIELD_KEYS,
    STUDENT_REQUIRED_IMPORT_KEYS,
    build_import_column_map,
    field_label_fr,
    normalize_mon_master_ranking,
    write_student_import_template,
    write_students_workbook,
)
from ..services.student_funding import encode_funding_codes, parse_funding_codes
from ..services.lookups import (
    gender_label_fr,
    is_valid_institutional_email,
    adapt_institutional_email,
    normalize_email,
    normalize_gender,
    normalize_level,
    normalize_track_acronym,
)
from ..services.student_status import (
    STUDENT_STATUS_GRADUATED,
    STUDENT_STATUS_WITHDRAWN,
    is_student_active,
    normalize_student_status,
)


_ENROLLMENT_FILTER_EMPTY = "__EMPTY_ENROLLMENT__"


class StudentsTab(QWidget):
    _TABLE_HEADERS = [
        "N° I.N.E.",
        "Nom",
        "Prénom",
        "Niveau",
        "Parcours",
        "Année",
        "Contrat péd.",
        "Email inst.",
    ]
    _ALARM_BG = QColor(255, 235, 238)
    _ALARM_FG = QColor(183, 28, 28)
    _WITHDRAWN_BG = QColor(245, 245, 245)
    _WITHDRAWN_FG = QColor(120, 120, 120)
    _GRADUATED_BG = QColor(232, 245, 233)
    _GRADUATED_FG = QColor(27, 94, 32)

    def __init__(self, repo, refresh_callbacks=None, default_academic_year: str = ""):
        super().__init__()
        self.repo = repo
        self.refresh_callbacks = refresh_callbacks or []
        self.default_academic_year = (default_academic_year or "").strip()
        self._students_raw: list[dict[str, Any]] = []
        self._visible_students: list[dict[str, Any]] = []
        self._missing_contract_ids: set[int] = set()
        layout = QVBoxLayout(self)
        layout.setSpacing(6)

        toolbar = QHBoxLayout()
        self.add_btn = QPushButton("Ajouter")
        self.add_btn.clicked.connect(self.add_student)
        self.profile_btn = QPushButton("Fiche…")
        self.profile_btn.clicked.connect(self.open_profile)
        self.edit_btn = QPushButton("Modifier…")
        self.edit_btn.clicked.connect(self.edit_student)
        self.withdraw_btn = QPushButton("Démissionnaire")
        self.withdraw_btn.setToolTip(
            "Retire l'étudiant de la liste active (notes, convocations, statistiques…) "
            "sans supprimer sa fiche ni son dossier de candidature."
        )
        self.withdraw_btn.clicked.connect(self._toggle_withdrawn_selection)
        toolbar.addWidget(self.add_btn)
        toolbar.addWidget(self.profile_btn)
        toolbar.addWidget(self.edit_btn)
        toolbar.addWidget(self.withdraw_btn)

        self.more_btn = QToolButton()
        self.more_btn.setText("Actions")
        self.more_btn.setPopupMode(QToolButton.ToolButtonPopupMode.InstantPopup)
        more_menu = QMenu(self)
        more_menu.addAction("Marquer démissionnaire", self.mark_selected_withdrawn)
        more_menu.addAction("Réintégrer (liste active)", self.restore_selected_active)
        more_menu.addSeparator()
        more_menu.addAction("Supprimer la sélection", self.delete_selected_students)
        more_menu.addSeparator()
        more_menu.addAction("Importer Excel…", self.import_excel)
        more_menu.addAction("Importer dossiers candidature (PDF)…", self.import_admission_pdfs)
        more_menu.addAction("Modèle d'import Excel…", self.generate_import_template)
        more_menu.addAction("Exporter Excel…", self.export_excel)
        more_menu.addSeparator()
        more_menu.addAction("Mailing liste (e-mail)…", self.open_email_list)
        more_menu.addSeparator()
        more_menu.addAction("Passage M2 / redoublement…", self.open_progression)
        more_menu.addAction(
            "Réinitialiser contrats pédagogiques M2 (passages antérieurs)…",
            self.reset_m2_pedagogical_contracts_batch,
        )
        self.more_btn.setMenu(more_menu)
        toolbar.addWidget(self.more_btn)
        toolbar.addStretch()
        layout.addLayout(toolbar)

        filters = QGridLayout()
        filters.setHorizontalSpacing(8)
        filters.addWidget(QLabel("Année :"), 0, 0)
        self.filter_year = QComboBox()
        self.filter_year.currentIndexChanged.connect(self._rebuild_table)
        filters.addWidget(self.filter_year, 0, 1)

        filters.addWidget(QLabel("Niveau :"), 0, 2)
        self.filter_level = QComboBox()
        self.filter_level.addItem("Tous", "")
        self.filter_level.addItem("M1", "M1")
        self.filter_level.addItem("M2", "M2")
        self.filter_level.currentIndexChanged.connect(self._on_filter_level_changed)
        filters.addWidget(self.filter_level, 0, 3)

        filters.addWidget(QLabel("Parcours :"), 1, 0)
        self.filter_track = QComboBox()
        self.filter_track.currentIndexChanged.connect(self._rebuild_table)
        filters.addWidget(self.filter_track, 1, 1)

        filters.addWidget(QLabel("Recherche :"), 1, 2)
        self.search_edit = QLineEdit()
        self.search_edit.setPlaceholderText("Nom, I.N.E., e-mail…")
        self.search_edit.textChanged.connect(self._rebuild_table)
        filters.addWidget(self.search_edit, 1, 3)

        filters.addWidget(QLabel("Contrat péd. :"), 2, 0)
        self.filter_contract = QComboBox()
        self.filter_contract.addItem("Tous", "all")
        self.filter_contract.addItem("⚠ Manquant (obligatoire)", "missing")
        self.filter_contract.addItem("Présent", "ok")
        self.filter_contract.currentIndexChanged.connect(self._rebuild_table)
        filters.addWidget(self.filter_contract, 2, 1)

        filters.addWidget(QLabel("Profil :"), 2, 2)
        self.filter_mobility = QComboBox()
        self.filter_mobility.addItem("Tous", "")
        self.filter_mobility.addItem("MNE (parcours complet)", "mne")
        self.filter_mobility.addItem("ERASMUS / mobilité", "erasmus")
        self.filter_mobility.currentIndexChanged.connect(self._rebuild_table)
        filters.addWidget(self.filter_mobility, 2, 3)

        filters.addWidget(QLabel("Affichage :"), 3, 0)
        self.filter_status = QComboBox()
        self.filter_status.addItem("Actifs", "active")
        self.filter_status.addItem("Diplômés", "graduated")
        self.filter_status.addItem("Démissionnaires", "withdrawn")
        self.filter_status.addItem("Tous", "all")
        self.filter_status.currentIndexChanged.connect(self._on_filter_status_changed)
        filters.addWidget(self.filter_status, 3, 1)

        filters.addWidget(QLabel("Établ. d'inscription :"), 3, 2)
        self.filter_enrollment = QComboBox()
        self.filter_enrollment.addItem("Tous", "")
        self.filter_enrollment.currentIndexChanged.connect(self._rebuild_table)
        filters.addWidget(self.filter_enrollment, 3, 3)

        sort_row = QHBoxLayout()
        sort_row.addWidget(QLabel("Tri :"))
        self.sort_combo = QComboBox()
        self.sort_combo.addItem("Nom (A → Z)", "last_name_asc")
        self.sort_combo.addItem("Nom (Z → A)", "last_name_desc")
        self.sort_combo.addItem("Prénom (A → Z)", "first_name_asc")
        self.sort_combo.currentIndexChanged.connect(self._rebuild_table)
        sort_row.addWidget(self.sort_combo)
        sort_row.addStretch()
        layout.addLayout(filters)
        layout.addLayout(sort_row)

        self.alarm_banner = QLabel()
        self.alarm_banner.setWordWrap(True)
        self.alarm_banner.hide()
        layout.addWidget(self.alarm_banner)

        hint = QLabel(
            "Double-cliquez sur une ligne pour ouvrir la fiche étudiant. "
            "« Démissionnaire » : masque l'étudiant des listes actives (notes, convocations…) "
            "tout en conservant sa fiche et son dossier."
        )
        hint.setWordWrap(True)
        hint.setStyleSheet("color: palette(mid); font-size: 11px;")
        layout.addWidget(hint)

        self.table = QTableWidget()
        self.table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.table.setSelectionMode(QAbstractItemView.SelectionMode.ExtendedSelection)
        self.table.doubleClicked.connect(self._on_double_click)
        layout.addWidget(self.table, 1)
        self._rebuild_track_filter()
        self._on_filter_status_changed()
        self.refresh()

    def _toggle_withdrawn_selection(self) -> None:
        mode = str(self.filter_status.currentData() or "active")
        if mode == "withdrawn":
            self.restore_selected_active()
        else:
            self.mark_selected_withdrawn()

    def mark_selected_withdrawn(self) -> None:
        ids = self._selected_student_ids()
        if not ids:
            QMessageBox.information(
                self,
                "Démissionnaire",
                "Sélectionnez un ou plusieurs étudiants dans la liste.",
            )
            return
        active_ids = [
            sid
            for sid in ids
            if is_student_active(self.repo.get_student(sid) or {})
        ]
        if not active_ids:
            QMessageBox.information(
                self,
                "Démissionnaire",
                "Les étudiants sélectionnés sont déjà marqués démissionnaires.",
            )
            return
        reply = QMessageBox.question(
            self,
            "Marquer démissionnaire",
            f"Marquer {len(active_ids)} étudiant(s) comme démissionnaire(s) ?\n\n"
            "Ils disparaîtront des listes actives (notes, convocations, statistiques…) "
            "mais leur fiche et leur dossier seront conservés.",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No,
        )
        if reply != QMessageBox.Yes:
            return
        try:
            self.repo.mark_students_withdrawn(active_ids)
            self.refresh()
            for cb in self.refresh_callbacks:
                cb()
        except Exception as exc:
            QMessageBox.critical(self, "Erreur", str(exc))

    def restore_selected_active(self) -> None:
        ids = self._selected_student_ids()
        if not ids:
            QMessageBox.information(
                self,
                "Réintégrer",
                "Sélectionnez un ou plusieurs étudiants dans la liste.",
            )
            return
        withdrawn_ids = [
            sid
            for sid in ids
            if not is_student_active(self.repo.get_student(sid) or {})
        ]
        if not withdrawn_ids:
            QMessageBox.information(
                self,
                "Réintégrer",
                "Les étudiants sélectionnés sont déjà actifs.",
            )
            return
        try:
            self.repo.restore_students_active(withdrawn_ids)
            self.refresh()
            for cb in self.refresh_callbacks:
                cb()
        except Exception as exc:
            QMessageBox.critical(self, "Erreur", str(exc))

    def _on_filter_level_changed(self) -> None:
        self._rebuild_track_filter()
        self._rebuild_table()

    def _rebuild_track_filter(self) -> None:
        prev = self.filter_track.currentData()
        level = str(self.filter_level.currentData() or "").strip().upper()
        self.filter_track.blockSignals(True)
        self.filter_track.clear()
        self.filter_track.addItem("Tous", "")
        if level in PARCOURS_BY_LEVEL:
            for code, _lab in PARCOURS_BY_LEVEL[level]:
                self.filter_track.addItem(track_display_label(level, code), code)
        else:
            seen: set[str] = set()
            for lv, lv_tracks in PARCOURS_BY_LEVEL.items():
                for code, _lab in lv_tracks:
                    if code not in seen:
                        seen.add(code)
                        self.filter_track.addItem(track_display_label(lv, code), code)
        if prev:
            idx = self.filter_track.findData(prev)
            if idx >= 0:
                self.filter_track.setCurrentIndex(idx)
        self.filter_track.blockSignals(False)

    def _on_filter_status_changed(self) -> None:
        mode = str(self.filter_status.currentData() or "active")
        if mode == "withdrawn":
            self.withdraw_btn.setText("Réintégrer")
            self.withdraw_btn.setToolTip(
                "Remet l'étudiant dans la liste active (notes, convocations, inscriptions…)."
            )
        else:
            self.withdraw_btn.setText("Démissionnaire")
            self.withdraw_btn.setToolTip(
                "Retire l'étudiant de la liste active sans supprimer sa fiche."
            )
        self._rebuild_table()

    def refresh(self) -> None:
        try:
            from ..services.student_parcours_repair import repair_student_parcours

            repair_student_parcours(self.repo.db)
        except Exception:
            pass
        self._students_raw = self.repo.list_students(include_withdrawn=True)
        self._missing_contract_ids = self.repo.student_ids_missing_pedagogical_contract()
        self._populate_year_filter()
        self._populate_enrollment_filter()
        self._rebuild_table()

    def _populate_enrollment_filter(self) -> None:
        from_db = {
            str(s.get("enrollment_institution") or "").strip()
            for s in self._students_raw
            if str(s.get("enrollment_institution") or "").strip()
        }
        known = set(MNE_ENROLLMENT_INSTITUTIONS)
        extra = sorted(from_db - known, key=str.casefold)
        has_empty = any(not str(s.get("enrollment_institution") or "").strip() for s in self._students_raw)

        prev = self.filter_enrollment.currentData()
        self.filter_enrollment.blockSignals(True)
        self.filter_enrollment.clear()
        self.filter_enrollment.addItem("Tous", "")
        for inst in MNE_ENROLLMENT_INSTITUTIONS:
            self.filter_enrollment.addItem(inst, inst)
        for inst in extra:
            self.filter_enrollment.addItem(inst, inst)
        if has_empty:
            self.filter_enrollment.addItem("(non renseigné)", _ENROLLMENT_FILTER_EMPTY)
        if prev is not None:
            idx = self.filter_enrollment.findData(prev)
            if idx >= 0:
                self.filter_enrollment.setCurrentIndex(idx)
            else:
                self.filter_enrollment.setCurrentIndex(0)
        else:
            self.filter_enrollment.setCurrentIndex(0)
        self.filter_enrollment.blockSignals(False)

    def _populate_year_filter(self) -> None:
        years = sorted(
            {str(s.get("academic_year") or "").strip() for s in self._students_raw if str(s.get("academic_year") or "").strip()},
            reverse=True,
        )
        prev = self.filter_year.currentData()
        self.filter_year.blockSignals(True)
        self.filter_year.clear()
        self.filter_year.addItem("Toutes", "")
        for y in years:
            self.filter_year.addItem(y, y)
        if prev:
            idx = self.filter_year.findData(prev)
            if idx >= 0:
                self.filter_year.setCurrentIndex(idx)
            elif self.default_academic_year:
                idx = self.filter_year.findData(self.default_academic_year)
                if idx < 0:
                    self.filter_year.addItem(self.default_academic_year, self.default_academic_year)
                    idx = self.filter_year.findData(self.default_academic_year)
                if idx >= 0:
                    self.filter_year.setCurrentIndex(idx)
            elif years:
                self.filter_year.setCurrentIndex(1)
            else:
                self.filter_year.setCurrentIndex(0)
        elif self.default_academic_year:
            idx = self.filter_year.findData(self.default_academic_year)
            if idx < 0:
                self.filter_year.addItem(self.default_academic_year, self.default_academic_year)
                idx = self.filter_year.findData(self.default_academic_year)
            if idx >= 0:
                self.filter_year.setCurrentIndex(idx)
            else:
                self.filter_year.setCurrentIndex(0)
        elif years:
            self.filter_year.setCurrentIndex(1)
        else:
            self.filter_year.setCurrentIndex(0)
        self.filter_year.blockSignals(False)

    def _selected_student_ids(self) -> list[int]:
        rows = sorted({idx.row() for idx in self.table.selectionModel().selectedRows()})
        ids: list[int] = []
        for row in rows:
            id_item = self.table.item(row, 0)
            if id_item is None:
                continue
            raw_id = id_item.data(Qt.ItemDataRole.UserRole)
            if raw_id is None:
                continue
            try:
                ids.append(int(raw_id))
            except (TypeError, ValueError):
                continue
        return ids

    def _selected_student_id(self) -> int | None:
        ids = self._selected_student_ids()
        if len(ids) != 1:
            return None
        return ids[0]

    def _selected_students(self) -> list[dict[str, Any]]:
        id_set = set(self._selected_student_ids())
        if not id_set:
            return []
        return [s for s in self._visible_students if int(s["id"]) in id_set]

    def open_email_list(self) -> None:
        if not self._visible_students:
            QMessageBox.information(
                self,
                "Liste d'e-mails",
                "Aucun étudiant ne correspond aux filtres actuels.",
            )
            return
        dlg = StudentEmailListDialog(
            filtered_students=self._visible_students,
            selected_students=self._selected_students(),
            parent=self,
        )
        dlg.exec()

    def _on_double_click(self, _index) -> None:
        self.open_profile()

    def open_profile(self) -> None:
        sid = self._selected_student_id()
        if sid is None:
            QMessageBox.information(self, "Fiche", "Sélectionnez un étudiant dans la liste.")
            return
        dlg = StudentProfileDialog(
            self.repo,
            sid,
            parent=self,
            default_academic_year=self.default_academic_year,
        )
        if dlg.exec():
            self.refresh()
            for cb in self.refresh_callbacks:
                cb()

    def _rebuild_table(self) -> None:
        data = list(self._students_raw)

        year_f = self.filter_year.currentData()
        if year_f:
            data = [s for s in data if str(s.get("academic_year") or "").strip() == year_f]

        track_f = self.filter_track.currentData()
        if track_f:
            data = [s for s in data if str(s.get("track") or "").strip().upper() == str(track_f).upper()]

        level_f = self.filter_level.currentData()
        if level_f:
            data = [s for s in data if str(s.get("level") or "").strip().upper() == str(level_f).upper()]

        q = self.search_edit.text().strip().lower()
        if q:

            def match(s: dict[str, Any]) -> bool:
                parts = [
                    s.get("student_number", ""),
                    s.get("student_number_ine", ""),
                    s.get("student_number_local", ""),
                    s.get("last_name", ""),
                    s.get("first_name", ""),
                    gender_label_fr(str(s.get("gender") or "")),
                    s.get("email_personal", ""),
                    s.get("email_institutional", ""),
                    s.get("nationality", ""),
                    s.get("birth_place", ""),
                ]
                return any(q in str(p).lower() for p in parts)

            data = [s for s in data if match(s)]

        contract_f = self.filter_contract.currentData()
        if contract_f == "missing":
            data = [s for s in data if int(s["id"]) in self._missing_contract_ids]
        elif contract_f == "ok":
            data = [s for s in data if int(s["id"]) not in self._missing_contract_ids]

        mobility_f = str(self.filter_mobility.currentData() or "")
        if mobility_f == "erasmus":
            data = [s for s in data if is_erasmus_student(s)]
        elif mobility_f == "mne":
            data = [s for s in data if not is_erasmus_student(s)]

        status_f = str(self.filter_status.currentData() or "active")
        if status_f == "active":
            data = [s for s in data if is_student_active(s)]
        elif status_f == "withdrawn":
            data = [
                s
                for s in data
                if normalize_student_status(s.get("status")) == STUDENT_STATUS_WITHDRAWN
            ]
        elif status_f == "graduated":
            data = [
                s
                for s in data
                if normalize_student_status(s.get("status")) == STUDENT_STATUS_GRADUATED
            ]

        inst_f = self.filter_enrollment.currentData()
        if inst_f == _ENROLLMENT_FILTER_EMPTY:
            data = [s for s in data if not str(s.get("enrollment_institution") or "").strip()]
        elif inst_f:
            data = [
                s
                for s in data
                if str(s.get("enrollment_institution") or "").strip() == str(inst_f)
            ]

        sort_key = self.sort_combo.currentData() or "last_name_asc"

        def last_name(s: dict[str, Any]) -> str:
            return str(s.get("last_name") or "").lower()

        def first_name(s: dict[str, Any]) -> str:
            return str(s.get("first_name") or "").lower()

        if sort_key == "last_name_asc":
            data.sort(key=lambda s: (last_name(s), first_name(s)))
        elif sort_key == "last_name_desc":
            data.sort(key=lambda s: (last_name(s), first_name(s)), reverse=True)
        elif sort_key == "first_name_asc":
            data.sort(key=lambda s: (first_name(s), last_name(s)))
        elif sort_key == "first_name_desc":
            data.sort(key=lambda s: (first_name(s), last_name(s)), reverse=True)

        self._fill_table(data)

    def _fill_table(self, students: list[dict[str, Any]]) -> None:
        self._visible_students = list(students)
        missing_visible = sum(1 for s in students if int(s["id"]) in self._missing_contract_ids)
        if missing_visible:
            self.alarm_banner.setText(
                f"⚠ {missing_visible} étudiant(s) affiché(s) sans contrat pédagogique signé "
                f"(document obligatoire). Filtrez avec « Contrat péd. → Manquant »."
            )
            self.alarm_banner.setStyleSheet(
                "background-color: #ffebee; color: #b71c1c; padding: 8px; "
                "border: 1px solid #ef9a9a; border-radius: 4px; font-weight: bold;"
            )
            self.alarm_banner.show()
        else:
            self.alarm_banner.hide()

        self.table.clear()
        self.table.setColumnCount(len(self._TABLE_HEADERS))
        self.table.setHorizontalHeaderLabels(self._TABLE_HEADERS)
        self.table.setRowCount(len(students))
        for r, s in enumerate(students):
            sid = int(s["id"])
            missing_contract = sid in self._missing_contract_ids
            status = normalize_student_status(s.get("status"))
            withdrawn = status == STUDENT_STATUS_WITHDRAWN
            graduated = status == STUDENT_STATUS_GRADUATED
            lv = str(s.get("level") or "")
            tr = str(s.get("track") or "")
            if is_erasmus_student(s):
                tr_disp = "ERASMUS"
                n_followed = len(
                    self.repo.list_student_erasmus_course_ids(
                        sid, str(s.get("academic_year") or "")
                    )
                )
                if n_followed:
                    tr_disp = f"ERASMUS ({n_followed} UE)"
            else:
                tr_disp = track_display_label(lv, tr)
            if sid in self._missing_contract_ids:
                contract_txt = "⚠ Manquant"
            else:
                paper = bool(int(s.get("pedagogical_contract_paper") or 0))
                has_pdf = self.repo.has_pedagogical_contract_pdf(sid)
                if paper and has_pdf:
                    contract_txt = "PDF + Papier"
                elif paper:
                    contract_txt = "Papier"
                else:
                    contract_txt = "PDF"
            vals = [
                s.get("student_number_ine", ""),
                s.get("last_name", ""),
                s.get("first_name", ""),
                lv,
                tr_disp,
                s.get("academic_year", ""),
                contract_txt,
                s.get("email_institutional", ""),
            ]
            for c, val in enumerate(vals):
                item = QTableWidgetItem(str(val))
                if c == 0:
                    item.setData(Qt.ItemDataRole.UserRole, sid)
                if missing_contract:
                    item.setBackground(QBrush(self._ALARM_BG))
                    if c == 6:
                        item.setForeground(QBrush(self._ALARM_FG))
                        f = item.font()
                        f.setBold(True)
                        item.setFont(f)
                elif graduated:
                    item.setBackground(QBrush(self._GRADUATED_BG))
                    item.setForeground(QBrush(self._GRADUATED_FG))
                elif withdrawn:
                    item.setBackground(QBrush(self._WITHDRAWN_BG))
                    item.setForeground(QBrush(self._WITHDRAWN_FG))
                    if c in (1, 2):
                        f = item.font()
                        f.setItalic(True)
                        item.setFont(f)
                self.table.setItem(r, c, item)
        self.table.resizeColumnsToContents()

    def generate_import_template(self) -> None:
        try:
            import openpyxl  # noqa: F401
        except Exception as exc:
            QMessageBox.critical(self, "Dépendance", f"openpyxl est requis.\n\n{exc}")
            return
        default_name = f"modele_import_etudiants_{date.today().isoformat()}.xlsx"
        path, _ = QFileDialog.getSaveFileName(
            self,
            "Générer modèle d'import étudiants",
            str(Path.home() / "Documents" / default_name),
            "Excel (*.xlsx)",
        )
        if not path:
            return
        if not path.lower().endswith(".xlsx"):
            path = path + ".xlsx"
        try:
            write_student_import_template(path)
            QMessageBox.information(
                self,
                "Modèle créé",
                f"Fichier enregistré :\n{path}\n\n"
                "Le numéro MNE n'est pas dans le modèle : il sera généré à l'import.",
            )
        except Exception as exc:
            QMessageBox.critical(self, "Erreur", str(exc))

    def export_excel(self) -> None:
        try:
            import openpyxl  # noqa: F401
        except Exception as exc:  # pragma: no cover
            QMessageBox.critical(self, "Dépendance", f"openpyxl est requis.\n\n{exc}")
            return

        n_all = len(self._students_raw)
        n_vis = len(self._visible_students)
        selected: list[dict[str, Any]] = []
        for row in sorted({idx.row() for idx in self.table.selectionModel().selectedRows()}):
            it = self.table.item(row, 0)
            if it is None:
                continue
            raw = it.data(Qt.ItemDataRole.UserRole)
            if raw is None:
                continue
            try:
                sid = int(raw)
            except (TypeError, ValueError):
                continue
            st = self.repo.get_student(sid)
            if st:
                selected.append(st)
        n_sel = len(selected)

        if n_all == 0:
            QMessageBox.information(self, "Export", "Aucun étudiant à exporter.")
            return

        dlg = QDialog(self)
        dlg.setWindowTitle("Exporter les étudiants")
        v = QVBoxLayout(dlg)
        grp = QButtonGroup(dlg)
        r_vis = QRadioButton(f"Liste filtrée — {n_vis} étudiant(s)")
        r_all = QRadioButton(f"Tous les étudiants — {n_all} étudiant(s)")
        r_sel = QRadioButton(f"Sélection uniquement — {n_sel} étudiant(s)")
        grp.addButton(r_vis)
        grp.addButton(r_all)
        grp.addButton(r_sel)
        r_vis.setChecked(True)
        r_sel.setEnabled(n_sel > 0)
        v.addWidget(r_vis)
        v.addWidget(r_all)
        v.addWidget(r_sel)
        buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        buttons.accepted.connect(dlg.accept)
        buttons.rejected.connect(dlg.reject)
        v.addWidget(buttons)
        if dlg.exec() != QDialog.DialogCode.Accepted:
            return

        if r_sel.isChecked():
            data = selected
        elif r_all.isChecked():
            data = list(self._students_raw)
        else:
            data = list(self._visible_students)

        if not data:
            QMessageBox.information(self, "Export", "Aucun étudiant pour ce choix.")
            return

        default_name = f"etudiants_{date.today().isoformat()}.xlsx"
        path, _ = QFileDialog.getSaveFileName(
            self,
            "Exporter les étudiants",
            str(Path.home() / "Documents" / default_name),
            "Excel (*.xlsx)",
        )
        if not path:
            return
        if not path.lower().endswith(".xlsx"):
            path = path + ".xlsx"

        try:
            write_students_workbook(path, data, include_instructions=True)
            QMessageBox.information(
                self,
                "Export terminé",
                f"{len(data)} étudiant(s) exporté(s) vers :\n{path}\n\n"
                f"L'export inclut l'identifiant interne ({len(STUDENT_EXPORT_FIELD_KEYS)} colonnes).",
            )
        except Exception as exc:
            QMessageBox.critical(self, "Erreur export", str(exc))

    def open_progression(self) -> None:
        sid = self._selected_student_id()
        if sid is None:
            QMessageBox.information(self, "Progression", "Sélectionnez exactement un étudiant.")
            return
        dlg = StudentProgressionDialog(
            self.repo,
            student_id=sid,
            default_academic_year=self.default_academic_year,
            parent=self,
        )
        if dlg.exec():
            self.refresh()
            for cb in self.refresh_callbacks:
                cb()

    def reset_m2_pedagogical_contracts_batch(self) -> None:
        candidates = self.repo.list_m2_students_with_pedagogical_contract()
        if not candidates:
            QMessageBox.information(
                self,
                "Contrats pédagogiques M2",
                "Aucun étudiant en M2 n'a encore de contrat pédagogique enregistré.",
            )
            return
        preview = "\n".join(
            f"  • {str(s.get('last_name') or '').strip()} {str(s.get('first_name') or '').strip()}".strip()
            for s in candidates[:15]
        )
        extra = ""
        if len(candidates) > 15:
            extra = f"\n  … et {len(candidates) - 15} autre(s)"
        reply = QMessageBox.question(
            self,
            "Contrats pédagogiques M2",
            (
                f"{len(candidates)} étudiant(s) en M2 ont encore un contrat pédagogique "
                "(probablement le contrat M1, distinct du contrat M2).\n\n"
                "Le contrat sera effacé sur chaque fiche (PDF + version papier). "
                "Un nouveau contrat M2 devra ensuite être déposé.\n\n"
                f"{preview}{extra}\n\n"
                "Confirmer la réinitialisation ?"
            ),
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.No,
        )
        if reply != QMessageBox.StandardButton.Yes:
            return
        count, names = self.repo.reset_pedagogical_contracts_for_m2_students()
        msg = f"{count} contrat(s) pédagogique(s) réinitialisé(s)."
        if names:
            lines = "\n".join(f"  • {n}" for n in names[:12])
            if len(names) > 12:
                lines += f"\n  … et {len(names) - 12} autre(s)"
            msg += f"\n\n{lines}"
        QMessageBox.information(self, "Contrats pédagogiques M2", msg)
        self.refresh()
        for cb in self.refresh_callbacks:
            cb()

    def add_student(self) -> None:
        dlg = StudentDialog(self, default_academic_year=self.default_academic_year, repo=self.repo)
        if dlg.exec():
            try:
                dlg.persist_create(self.repo)
                self.refresh()
                for cb in self.refresh_callbacks:
                    cb()
            except Exception as exc:
                QMessageBox.critical(self, "Erreur", str(exc))

    def edit_student(self) -> None:
        sid = self._selected_student_id()
        if sid is None:
            QMessageBox.information(self, "Modifier", "Sélectionnez un étudiant.")
            return
        student = self.repo.get_student(sid)
        if student is None:
            QMessageBox.warning(self, "Modifier", "Étudiant introuvable.")
            return
        dlg = StudentDialog(self, student=student, repo=self.repo)
        if dlg.exec():
            try:
                sn = str(student.get("student_number") or "")
                dlg.persist_update(self.repo, sid, sn)
                self.refresh()
                for cb in self.refresh_callbacks:
                    cb()
            except Exception as exc:
                QMessageBox.critical(self, "Erreur", str(exc))

    def delete_selected_students(self) -> None:
        rows = sorted({idx.row() for idx in self.table.selectionModel().selectedRows()})
        if not rows:
            QMessageBox.information(self, "Supprimer", "Sélectionnez une ou plusieurs lignes.")
            return

        ids: list[int] = []
        labels: list[str] = []
        for row in rows:
            id_item = self.table.item(row, 0)
            if id_item is None:
                continue
            student_id = id_item.data(Qt.ItemDataRole.UserRole)
            if student_id is None:
                continue
            try:
                sid = int(student_id)
            except (TypeError, ValueError):
                continue
            ids.append(sid)
            num_item = self.table.item(row, 0)
            name_item = self.table.item(row, 1)
            label = f"#{sid}"
            if num_item and num_item.text():
                label = num_item.text()
            if name_item and name_item.text():
                label = f"{label} — {name_item.text()}"
            labels.append(label)

        if not ids:
            return

        if len(ids) == 1:
            body = f"Supprimer {labels[0]} ?\nLes inscriptions et notes associées seront effacées."
        else:
            preview = "\n".join(labels[:10])
            if len(labels) > 10:
                preview += f"\n… et {len(labels) - 10} de plus"
            body = f"Supprimer {len(ids)} étudiants ?\n\n{preview}"

        reply = QMessageBox.question(
            self,
            "Confirmer",
            body,
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No,
        )
        if reply != QMessageBox.Yes:
            return
        try:
            self.repo.delete_students(ids)
            self.refresh()
            for cb in self.refresh_callbacks:
                cb()
        except Exception as exc:
            QMessageBox.critical(self, "Erreur", str(exc))

    def import_admission_pdfs(self) -> None:
        paths, _ = QFileDialog.getOpenFileNames(
            self,
            "Importer des dossiers de candidature",
            "",
            "PDF (*.pdf)",
        )
        if not paths:
            folder = QFileDialog.getExistingDirectory(
                self,
                "Ou choisir un dossier de candidatures",
            )
            if not folder:
                return
            paths = [folder]

        pdf_files = collect_admission_pdfs(paths)
        if not pdf_files:
            QMessageBox.information(self, "Import", "Aucun fichier PDF trouvé.")
            return

        dossiers = [parse_admission_pdf(p) for p in pdf_files]
        dlg = AdmissionImportDialog(
            dossiers,
            repo=self.repo,
            default_academic_year=self.default_academic_year,
            parent=self,
        )
        if dlg.exec() != QDialog.DialogCode.Accepted:
            return

        selected = dlg.selected_dossiers()
        created, updated, skipped, errors = import_admission_dossiers(
            self.repo,
            selected,
            default_academic_year=self.default_academic_year,
            update_existing=dlg.should_update_existing(),
        )
        self.refresh()
        for cb in self.refresh_callbacks:
            cb()

        msg = f"Créés : {created}\nMis à jour : {updated}\nIgnorés : {skipped}"
        if errors:
            preview = "\n".join(errors[:15])
            if len(errors) > 15:
                preview += f"\n… ({len(errors) - 15} de plus)"
            msg += f"\n\nDétails :\n{preview}"
        QMessageBox.information(self, "Import candidatures", msg)

    def import_excel(self) -> None:
        path, _ = QFileDialog.getOpenFileName(
            self,
            "Importer des étudiants",
            "",
            "Excel (*.xlsx)",
        )
        if not path:
            return

        try:
            from openpyxl import load_workbook
        except Exception as exc:  # pragma: no cover
            QMessageBox.critical(self, "Dépendance", f"openpyxl est requis.\n\n{exc}")
            return

        try:
            wb = load_workbook(path, data_only=True)
            ws = wb.active
            rows_iter = ws.iter_rows(values_only=True)
            header_row = next(rows_iter, None)
            if not header_row:
                QMessageBox.warning(self, "Import", "Fichier vide.")
                return

            col_map = build_import_column_map(header_row)
            col_student_number_ine = col_map.get("student_number_ine")
            col_student_number_local = col_map.get("student_number_local")
            col_last_name = col_map.get("last_name")
            col_first_name = col_map.get("first_name")
            col_email_personal = col_map.get("email_personal")
            col_email_institutional = col_map.get("email_institutional")
            col_phone = col_map.get("phone")
            col_enrollment_institution = col_map.get("enrollment_institution")
            col_application_platform = col_map.get("application_platform")
            col_mon_master_ranking = col_map.get("mon_master_ranking")
            col_funding = col_map.get("funding")
            col_funding_other = col_map.get("funding_other")
            col_accommodations = col_map.get("accommodations")
            col_accommodations_other = col_map.get("accommodations_other")
            col_notes = col_map.get("notes")
            col_level = col_map.get("level")
            col_track = col_map.get("track")
            col_academic_year = col_map.get("academic_year")
            col_birth_date = col_map.get("birth_date")
            col_nationality = col_map.get("nationality")
            col_origin_institution = col_map.get("origin_institution")
            col_origin_country = col_map.get("origin_institution_country")
            col_highest_diploma = col_map.get("highest_diploma")
            col_birth_place = col_map.get("birth_place")
            col_gender = col_map.get("gender")
            col_flag_m1c = col_map.get("m1_c")
            col_flag_m1p = col_map.get("m1p")

            missing = [k for k in STUDENT_REQUIRED_IMPORT_KEYS if col_map.get(k) is None]
            if missing:
                QMessageBox.critical(
                    self,
                    "Import",
                    "Colonnes obligatoires manquantes : "
                    + ", ".join(field_label_fr(k) for k in missing)
                    + "\n\nL'identifiant interne MNE est généré par l'application. "
                    "Utilisez « Modèle d'import Excel… ».",
                )
                return

            created = 0
            skipped = 0
            errors: list[str] = []

            def cell(row: tuple[Any, ...], idx: int | None) -> str:
                if idx is None or idx >= len(row):
                    return ""
                v = row[idx]
                return "" if v is None else str(v).strip()

            def cell_raw(row: tuple[Any, ...], idx: int | None) -> Any:
                if idx is None or idx >= len(row):
                    return None
                return row[idx]

            for excel_row_idx, row in enumerate(rows_iter, start=2):
                student_number_ine = cell(row, col_student_number_ine)
                student_number_local = cell(row, col_student_number_local)
                last_name = cell(row, col_last_name)
                first_name = cell(row, col_first_name)
                if not (last_name and first_name):
                    skipped += 1
                    continue

                email_personal = normalize_email(cell(row, col_email_personal))
                enrollment_institution = cell(row, col_enrollment_institution)
                email_institutional = adapt_institutional_email(
                    first_name,
                    last_name,
                    enrollment_institution,
                    normalize_email(cell(row, col_email_institutional)),
                )
                phone = cell(row, col_phone)
                application_platform = cell(row, col_application_platform)
                mon_master_ranking = normalize_mon_master_ranking(
                    cell_raw(row, col_mon_master_ranking)
                )
                funding = encode_funding_codes(parse_funding_codes(cell(row, col_funding)))
                funding_other = cell(row, col_funding_other)
                accommodations = cell(row, col_accommodations)
                accommodations_other = cell(row, col_accommodations_other)
                notes = cell(row, col_notes)
                level = normalize_level(cell(row, col_level))
                track = normalize_track_acronym(cell(row, col_track))
                academic_year = (
                    self.default_academic_year
                    if self.default_academic_year
                    else cell(row, col_academic_year)
                )

                birth_date = normalize_birth_date_iso(cell_raw(row, col_birth_date))
                nationality = cell(row, col_nationality)
                origin_institution = cell(row, col_origin_institution)
                origin_institution_country = cell(row, col_origin_country)
                highest_diploma = cell(row, col_highest_diploma)
                birth_place = cell(row, col_birth_place)
                gender = normalize_gender(cell_raw(row, col_gender))

                if not track:

                    def truthy(v: Any) -> bool:
                        if v is None:
                            return False
                        if isinstance(v, bool):
                            return v
                        s = str(v).strip().lower()
                        return s in {"1", "x", "yes", "y", "true", "v", "ok"}

                    if col_flag_m1c is not None and col_flag_m1c < len(row) and truthy(row[col_flag_m1c]):
                        track = "C"
                    elif col_flag_m1p is not None and col_flag_m1p < len(row) and truthy(row[col_flag_m1p]):
                        track = "P"

                if not level and track:
                    if track in {"P", "C", "M1P", "M1C"}:
                        level = "M1"
                    elif track in {"NFC", "DWM", "NPO", "NPD", "NRPE"}:
                        level = "M2"

                if not is_valid_institutional_email(email_institutional):
                    errors.append(
                        f"Ligne {excel_row_idx} : email institutionnel invalide « {email_institutional} »."
                    )
                    continue
                if level and level not in {"M1", "M2"}:
                    errors.append(f"Ligne {excel_row_idx} : niveau invalide « {level} ».")
                    continue
                if track:
                    from ..services.lookups import TRACKS

                    if track not in TRACKS:
                        errors.append(f"Ligne {excel_row_idx} : parcours invalide « {track} ».")
                        continue
                if col_gender is not None and cell(row, col_gender) and not gender:
                    errors.append(f"Ligne {excel_row_idx} : genre non reconnu.")
                    continue

                try:
                    new_id = self.repo.add_student(
                        "",
                        student_number_ine,
                        student_number_local,
                        last_name,
                        first_name,
                        email_personal,
                        email_institutional,
                        phone,
                        enrollment_institution,
                        application_platform,
                        mon_master_ranking,
                        accommodations,
                        accommodations_other,
                        funding,
                        funding_other,
                        notes,
                        level,
                        track,
                        academic_year,
                        birth_date=birth_date,
                        nationality=nationality,
                        birth_place=birth_place,
                        gender=gender,
                        origin_institution=origin_institution,
                        origin_institution_country=origin_institution_country,
                        highest_diploma=highest_diploma,
                    )
                    self.repo.sync_enrollments_for_student(new_id)
                    created += 1
                except Exception as exc:
                    errors.append(f"Ligne {excel_row_idx} : {exc}")

            self.refresh()
            for cb in self.refresh_callbacks:
                cb()

            msg = f"Importés : {created}\nIgnorés (ligne incomplète) : {skipped}"
            if errors:
                preview = "\n".join(errors[:15])
                if len(errors) > 15:
                    preview += f"\n… ({len(errors) - 15} de plus)"
                msg += f"\n\nErreurs :\n{preview}"
            QMessageBox.information(self, "Import terminé", msg)

        except Exception as exc:
            QMessageBox.critical(self, "Erreur import", str(exc))
