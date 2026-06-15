from __future__ import annotations

import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from PySide6.QtCore import QEvent, Qt
from PySide6.QtGui import QKeySequence
from PySide6.QtWidgets import (
    QDialog,
    QCheckBox,
    QComboBox,
    QFileDialog,
    QHBoxLayout,
    QLabel,
    QMessageBox,
    QPushButton,
    QTableWidget,
    QTableWidgetItem,
    QVBoxLayout,
    QWidget,
)

from ..core.parcours import track_label
from ..services.grade_status import format_grade_display, parse_grade_cell
from ..services.grades_excel import write_grades_import_template, write_grades_workbook


@dataclass(frozen=True)
class _ColMap:
    session: int
    kind: str
    coefficient: float
    assessment_id: int
    header: str


def _norm_number(s: Any) -> str:
    if s is None:
        return ""
    return str(s).replace("\xa0", " ").strip()


def _parse_assessment_header(text: str) -> tuple[str, float] | None:
    """
    Parse headers like:
    - "EE (0,4)"
    - "EEF (0,6)"
    - "CC (0,3)"
    - "EE (1)"
    """
    t = (text or "").strip()
    m = re.match(r"^(?P<kind>[A-Za-zÀ-ÖØ-öø-ÿ_]{2,12})\s*\(\s*(?P<coef>[0-9]+(?:[.,][0-9]+)?)\s*\)\s*$", t)
    if not m:
        return None
    kind = m.group("kind").strip().upper()
    coef = float(m.group("coef").replace(",", "."))
    return kind, coef


class CourseGradesDialog(QDialog):
    """
    Saisie "matière par matière" : tableau étudiants x évaluations (assessments).
    Supporte coller depuis Excel (TSV) et import d'un fichier de notes M1NE.
    """

    def __init__(self, repo, *, template_id: int, course_id: int, parent=None):
        super().__init__(parent)
        self.repo = repo
        self.template_id = int(template_id)
        self.course_id = int(course_id)
        self.setWindowTitle("Saisie des notes — par matière")

        layout = QVBoxLayout(self)
        self.info = QLabel("")
        self.info.setWordWrap(True)
        layout.addWidget(self.info)

        actions = QHBoxLayout()
        self.scope_common = QCheckBox(
            "UE commune : tous les parcours inscrits (fusion des maquettes)"
        )
        self.scope_common.setToolTip(
            "Regroupe les étudiants inscrits à toutes les maquettes qui contiennent cette UE "
            "(ex. M1 P + M1 C), sans repasser par chaque parcours."
        )
        self.scope_common.toggled.connect(lambda _checked: self.refresh())
        self.entry_session = QComboBox()
        self.entry_session.addItem("Saisie : Session 1", 1)
        self.entry_session.addItem("Saisie : Session 2", 2)
        self.entry_session.currentIndexChanged.connect(lambda _idx: self.refresh())
        self.export_template_btn = QPushButton("Générer modèle Excel…")
        self.export_template_btn.clicked.connect(self.export_template_xlsx)
        self.export_grades_btn = QPushButton("Exporter notes (Excel)…")
        self.export_grades_btn.clicked.connect(self.export_grades_xlsx)
        self.import_btn = QPushButton("Importer un fichier de notes Excel (.xlsx)…")
        self.import_btn.clicked.connect(self.import_xlsx)
        self.save_btn = QPushButton("Enregistrer dans la base")
        self.save_btn.clicked.connect(self.save_all)
        actions.addWidget(self.scope_common)
        actions.addWidget(self.entry_session)
        actions.addWidget(self.export_template_btn)
        actions.addWidget(self.export_grades_btn)
        actions.addWidget(self.import_btn)
        actions.addWidget(self.save_btn)
        actions.addStretch()
        layout.addLayout(actions)

        self.table = QTableWidget()
        self.table.setSelectionMode(QTableWidget.SelectionMode.ExtendedSelection)
        self.table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectItems)
        self.table.installEventFilter(self)
        layout.addWidget(self.table)

        self._student_rows: list[dict[str, Any]] = []
        self._col_maps: list[_ColMap] = []
        self._existing_locked: dict[int, dict[int, int]] = {}
        self._validated_col = 3
        self._first_grade_col = 3
        self._show_parcours_col = False
        self._student_template_by_id: dict[int, int] = {}
        self._scope_academic_year = ""

        if self._course_is_multi_parcours():
            self.scope_common.setChecked(True)

        self.refresh()

        from .screen_layout import adapt_window_size

        adapt_window_size(self, preferred=(1200, 720), minimum=(720, 480))

    def _current_template_row(self) -> dict[str, Any]:
        return (
            next(
                (x for x in self.repo.list_templates() if int(x["id"]) == self.template_id),
                None,
            )
            or {}
        )

    def _resolve_scope_academic_year(self) -> str:
        ay = str(self._current_template_row().get("academic_year") or "").strip()
        if ay:
            return ay
        inferred = self.repo.infer_academic_years_for_course(self.course_id)
        if len(inferred) == 1:
            return inferred[0]
        return ""

    def _course_template_ids(self, academic_year: str = "") -> list[int]:
        return self.repo.list_template_ids_with_course(
            self.course_id, academic_year=academic_year
        )

    def _course_is_multi_parcours(self) -> bool:
        ay = self._resolve_scope_academic_year()
        ids = self._course_template_ids(ay) if ay else self._course_template_ids()
        return len(ids) > 1

    def _template_for_student(self, student_id: int) -> int:
        return int(self._student_template_by_id.get(int(student_id), self.template_id))

    def _parcours_label(self, student: dict[str, Any]) -> str:
        lv = str(student.get("level") or "").strip()
        tr = str(student.get("track") or "").strip()
        if not tr:
            return ""
        lab = track_label(lv, tr)
        return f"{lab} ({tr})" if lab and lab != tr else tr

    def _load_student_scope(self) -> str:
        ay = self._resolve_scope_academic_year()
        self._scope_academic_year = ay
        if self.scope_common.isChecked():
            template_ids = self._course_template_ids(ay) if ay else self._course_template_ids()
            if len(template_ids) > 1:
                templates = self.repo.list_templates_containing_course(
                    self.course_id, academic_year=ay
                )
                parcours_bits = sorted(
                    {
                        f"{str(t.get('level') or '').strip()} {str(t.get('track') or '').strip()}".strip()
                        for t in templates
                        if str(t.get("track") or "").strip()
                    }
                )
                parcours_txt = ", ".join(parcours_bits) if parcours_bits else "—"
                self._student_rows = self.repo.list_students_for_course_in_templates(
                    self.course_id, template_ids
                )
                self._student_template_by_id = self.repo.student_templates_for_course(
                    self.course_id, template_ids
                )
                self._show_parcours_col = True
                year_txt = f" — {ay}" if ay else ""
                return (
                    f"UE commune : {len(template_ids)} maquette(s) "
                    f"({parcours_txt}){year_txt}"
                )
            if template_ids:
                self._student_rows = self.repo.list_students_for_course_in_templates(
                    self.course_id, template_ids
                )
                self._student_template_by_id = self.repo.student_templates_for_course(
                    self.course_id, template_ids
                )
                self._show_parcours_col = False
                return "UE commune : une seule maquette contient cette UE"

        self._student_rows = self.repo.list_students_for_template(self.template_id)
        self._student_template_by_id = {
            int(s["id"]): self.template_id for s in self._student_rows
        }
        self._show_parcours_col = False
        t_row = self._current_template_row()
        lv = str(t_row.get("level") or "").strip()
        tr = str(t_row.get("track") or "").strip()
        suffix = f" ({lv} {tr})".strip() if lv or tr else ""
        return f"Maquette courante{suffix}"

    def refresh(self) -> None:
        course = self.repo.get_course(self.course_id) or {}
        title = f"{course.get('code','')} — {course.get('name','')}".strip(" —")
        scope_txt = self._load_student_scope()
        self._first_grade_col = 4 if self._show_parcours_col else 3

        self.info.setText(
            f"<b>{title}</b><br/>"
            f"<i>Périmètre</i> : {scope_txt} — <b>{len(self._student_rows)}</b> étudiant(s)<br/>"
            "Collez un tableau depuis Excel (Ctrl/Cmd+V) ou importez un .xlsx. "
            "Sélectionnez des cellules de notes puis <b>Suppr</b> ou <b>Retour arrière</b> pour les vider "
            "(sauf notes verrouillées dans l’onglet Notes). "
            "Valeurs : nombre, ABJ, DEF, NEUT, VAL — ou cochez « Validée » pour valider l’UE "
            "sans note chiffrée (UE libre, équivalence, etc.)."
        )
        assessments = self.repo.list_assessments(self.course_id)

        # Build column maps: one column per assessment (session+kind+coef)
        self._col_maps = []
        headers = ["N° I.N.E.", "Nom", "Prénom"]
        if self._show_parcours_col:
            headers.append("Parcours")
        for a in assessments:
            headers.append(f"S{a['session']} {a['kind']} ({float(a['coefficient']):g})")
            self._col_maps.append(
                _ColMap(
                    session=int(a["session"]),
                    kind=str(a["kind"]).strip().upper(),
                    coefficient=float(a["coefficient"]),
                    assessment_id=int(a["id"]),
                    header=str(headers[-1]),
                )
            )
        headers.append("Validée")
        self._validated_col = len(headers) - 1

        self.table.clear()
        self.table.setColumnCount(len(headers))
        self.table.setHorizontalHeaderLabels(headers)
        self.table.setRowCount(len(self._student_rows))

        # Fill base identity + existing grades for each assessment
        from ..services.lookups import student_transcript_number

        grade_col = self._first_grade_col
        for r, s in enumerate(self._student_rows):
            self.table.setItem(r, 0, QTableWidgetItem(student_transcript_number(s)))
            self.table.setItem(r, 1, QTableWidgetItem(str(s.get("last_name") or "")))
            self.table.setItem(r, 2, QTableWidgetItem(str(s.get("first_name") or "")))
            if self._show_parcours_col:
                par_it = QTableWidgetItem(self._parcours_label(s))
                par_it.setFlags(par_it.flags() & ~Qt.ItemFlag.ItemIsEditable)
                self.table.setItem(r, 3, par_it)

        # Bulk load existing grades
        self._existing_locked = {}
        for r, s in enumerate(self._student_rows):
            sid = int(s["id"])
            tid = self._template_for_student(sid)
            if self.repo.is_sent_to_second_session(sid, tid, self.course_id):
                self.repo.carry_over_reprise_grades_from_session1(sid, self.course_id)
            rows = self.repo.get_grades_for_student_course(sid, self.course_id)
            by_assessment = {int(x["assessment_id"]): x for x in rows}
            self._existing_locked[int(s["id"])] = {
                int(aid): int(v.get("locked") or 0) for aid, v in by_assessment.items()
            }
            for c, cm in enumerate(self._col_maps, start=grade_col):
                g = by_assessment.get(int(cm.assessment_id))
                if g is None:
                    txt = ""
                else:
                    txt = format_grade_display(
                        g.get("grade"),
                        g.get("status"),
                        assessment_session=int(g.get("session") or 1),
                    )
                it = QTableWidgetItem(txt)
                self.table.setItem(r, c, it)

            validated = self.repo.has_ue_ects_validation(sid, tid, self.course_id)
            val_it = QTableWidgetItem("")
            val_it.setFlags(
                val_it.flags() | Qt.ItemFlag.ItemIsUserCheckable | Qt.ItemFlag.ItemIsEnabled
            )
            val_it.setCheckState(
                Qt.CheckState.Checked if validated else Qt.CheckState.Unchecked
            )
            self.table.setItem(r, self._validated_col, val_it)
            self._apply_validated_row_style(r, validated)

        self.table.resizeColumnsToContents()

    def _apply_validated_row_style(self, row: int, validated: bool) -> None:
        """Grise les notes si l’UE est validée sans moyenne chiffrée."""
        first_grade_col = self._first_grade_col
        last_grade_col = self._validated_col
        for c in range(first_grade_col, last_grade_col):
            it = self.table.item(row, c)
            if it is None:
                continue
            if validated:
                it.setFlags(it.flags() & ~Qt.ItemFlag.ItemIsEditable)
            else:
                it.setFlags(it.flags() | Qt.ItemFlag.ItemIsEditable)

    def eventFilter(self, obj, event):  # noqa: ANN001
        if obj is self.table and event.type() == QEvent.Type.KeyPress:
            if event.key() in (Qt.Key.Key_Delete, Qt.Key.Key_Backspace):
                if self._clear_selected_grade_cells():
                    return True
        return super().eventFilter(obj, event)

    def _clear_selected_grade_cells(self) -> bool:
        """Efface le texte des cellules d’évaluation sélectionnées (col. ≥ 3), sauf si verrouillées."""
        idxs = self.table.selectedIndexes()
        if not idxs or not self._col_maps:
            return False
        first_assessment_col = self._first_grade_col
        done = False
        for ix in idxs:
            c = ix.column()
            if c < first_assessment_col or c >= getattr(self, "_validated_col", self.table.columnCount()):
                continue
            ci = c - first_assessment_col
            if ci < 0 or ci >= len(self._col_maps):
                continue
            r = ix.row()
            if r < 0 or r >= len(self._student_rows):
                continue
            sid = int(self._student_rows[r]["id"])
            cm = self._col_maps[ci]
            locked = int((self._existing_locked.get(sid) or {}).get(int(cm.assessment_id)) or 0)
            if locked:
                continue
            it = self.table.item(r, c)
            if it is None:
                continue
            it.setText("")
            done = True
        return done

    def keyPressEvent(self, event):  # noqa: N802
        if event.matches(QKeySequence.Paste):
            self.paste_tsv()
            return
        super().keyPressEvent(event)

    def paste_tsv(self) -> None:
        cb = self.table.clipboard() if hasattr(self.table, "clipboard") else None
        # Qt doesn't expose clipboard on widget; use QApplication
        from PySide6.QtWidgets import QApplication

        text = QApplication.clipboard().text()
        if not text.strip():
            return
        # IMPORTANT: preserve empty lines.
        # If we drop blank lines from the clipboard, all subsequent rows shift up,
        # which looks like the app "invented" grades.
        lines = text.splitlines()
        raw_rows = [line.split("\t") for line in lines]
        if not raw_rows:
            return
        max_cols = max((len(r) for r in raw_rows), default=0)
        if max_cols <= 0:
            return
        # Normalize row widths so a blank line advances the row cursor for all columns.
        raw_rows = [(r + [""] * (max_cols - len(r))) for r in raw_rows]

        # Mode robuste : si la première colonne ressemble à des numéros étudiants présents,
        # on colle "par numéro" (évite les décalages et les valeurs collées sur le mauvais étudiant).
        by_num: dict[str, int] = {}
        for r in range(self.table.rowCount()):
            it = self.table.item(r, 0)
            num = (it.text().strip() if it else "").strip()
            if num:
                by_num[num] = r

        def is_header_cell(v: str) -> bool:
            vv = (v or "").strip().lower()
            return vv in {"numéro", "numero", "student #", "student#", "ine", "n°"}

        start_idx = 1 if raw_rows and raw_rows[0] and is_header_cell(str(raw_rows[0][0])) else 0
        data_rows = raw_rows[start_idx:]

        # Heuristique : au moins 3 lignes matchent un numéro connu => collage par numéro.
        hit = 0
        for row in data_rows[:50]:
            if row and str(row[0]).strip() in by_num:
                hit += 1
        paste_by_number = hit >= 3

        if paste_by_number:
            # Colonnes attendues : Numéro | (Nom) | (Prénom) | puis évaluations...
            paste_start = self._first_grade_col
            for row in data_rows:
                if not row:
                    continue
                stu_num = str(row[0]).strip()
                rr = by_num.get(stu_num)
                if rr is None:
                    continue
                for i, val in enumerate(row[paste_start:], start=paste_start):
                    cc = i
                    if cc >= self.table.columnCount():
                        break
                    self.table.setItem(rr, cc, QTableWidgetItem(str(val).strip()))
            return

        # Fallback : collage "grille" classique depuis la cellule courante
        # Prefer selection top-left (more reliable than currentRow for users).
        r0, c0 = self.table.currentRow(), self.table.currentColumn()
        sel = self.table.selectedRanges()
        if sel:
            r0 = sel[0].topRow()
            c0 = sel[0].leftColumn()
        if r0 < 0 or c0 < 0:
            r0, c0 = 0, self._first_grade_col

        # If user selected rows and clipboard rows count matches, map by selected rows order.
        selected_rows = sorted({idx.row() for idx in self.table.selectionModel().selectedRows()})
        if selected_rows and len(selected_rows) == len(raw_rows) and c0 >= self._first_grade_col:
            for dr, row in enumerate(raw_rows):
                rr = selected_rows[dr]
                for dc, val in enumerate(row):
                    cc = c0 + dc
                    if cc >= self.table.columnCount():
                        break
                    self.table.setItem(rr, cc, QTableWidgetItem(str(val).strip()))
            return

        for dr, row in enumerate(raw_rows):
            for dc, val in enumerate(row):
                rr = r0 + dr
                cc = c0 + dc
                if rr >= self.table.rowCount() or cc >= self.table.columnCount():
                    continue
                self.table.setItem(rr, cc, QTableWidgetItem(str(val).strip()))

    def save_all(self) -> None:
        # Save all grade cells (assessments columns only)
        try:
            target_session = int(self.entry_session.currentData() or 1)
            for r, s in enumerate(self._student_rows):
                sid = int(s["id"])
                tid = self._template_for_student(sid)
                val_it = self.table.item(r, self._validated_col)
                validated = bool(
                    val_it and val_it.checkState() == Qt.CheckState.Checked
                )
                self.repo.set_ue_ects_validation(
                    sid,
                    tid,
                    self.course_id,
                    validated=validated,
                )
                self._apply_validated_row_style(r, validated)
                if validated:
                    continue
                for c, cm in enumerate(self._col_maps, start=self._first_grade_col):
                    # Ne pas écrire sur l'autre session, sinon on force une S2 "à 0" en mettant DEF partout.
                    if int(cm.session) != target_session:
                        continue
                    it = self.table.item(r, c)
                    raw = it.text().strip() if it else ""
                    locked = int((self._existing_locked.get(sid) or {}).get(int(cm.assessment_id)) or 0)
                    if raw == "":
                        if locked:
                            continue
                        self.repo.upsert_grade(
                            sid,
                            int(cm.assessment_id),
                            None,
                            status="OK",
                            locked=0,
                            comment="",
                        )
                        continue
                    grade, status, err = parse_grade_cell(raw)
                    if err:
                        QMessageBox.warning(self, "Grade", f"Row {r+1}, {cm.header}: {err}")
                        return
                    self.repo.upsert_grade(
                        sid,
                        int(cm.assessment_id),
                        grade,
                        status=status,
                        locked=locked,
                        comment="",
                    )
            QMessageBox.information(self, "Enregistrement", "Notes enregistrées.")
        except Exception as exc:
            QMessageBox.critical(self, "Erreur", str(exc))

    def _grades_export_path(self, suffix: str) -> str | None:
        course = self.repo.get_course(self.course_id) or {}
        code = str(course.get("code") or "notes").replace("/", "-")
        default = str(Path.home() / "Documents" / f"{code}_{suffix}.xlsx")
        path, _ = QFileDialog.getSaveFileName(
            self,
            "Enregistrer le fichier Excel",
            default,
            "Excel (*.xlsx)",
        )
        if not path:
            return None
        if not path.lower().endswith(".xlsx"):
            path = path + ".xlsx"
        return path

    def export_template_xlsx(self) -> None:
        try:
            import openpyxl  # noqa: F401
        except Exception as exc:
            QMessageBox.critical(self, "Dépendance", f"openpyxl est requis.\n\n{exc}")
            return
        if not self._student_rows:
            QMessageBox.information(self, "Export", "Aucun étudiant dans le périmètre sélectionné.")
            return
        path = self._grades_export_path("modele_notes")
        if not path:
            return
        try:
            write_grades_import_template(
                self.repo,
                template_id=self.template_id,
                course_id=self.course_id,
                path=path,
                students=self._student_rows,
            )
            QMessageBox.information(
                self,
                "Modèle créé",
                f"Fichier enregistré :\n{path}\n\n"
                "Remplissez les notes puis importez via « Importer un fichier de notes Excel ».",
            )
        except Exception as exc:
            QMessageBox.critical(self, "Export", str(exc))

    def export_grades_xlsx(self) -> None:
        try:
            import openpyxl  # noqa: F401
        except Exception as exc:
            QMessageBox.critical(self, "Dépendance", f"openpyxl est requis.\n\n{exc}")
            return
        if not self._student_rows:
            QMessageBox.information(self, "Export", "Aucun étudiant dans le périmètre sélectionné.")
            return
        path = self._grades_export_path("notes")
        if not path:
            return
        try:
            write_grades_workbook(
                self.repo,
                template_id=self.template_id,
                course_id=self.course_id,
                path=path,
                students=self._student_rows,
                fill_grades=True,
                include_instructions=True,
            )
            QMessageBox.information(
                self,
                "Export",
                f"Notes exportées ({len(self._student_rows)} étudiant(s)) :\n{path}",
            )
        except Exception as exc:
            QMessageBox.critical(self, "Export", str(exc))

    def import_xlsx(self) -> None:
        """
        Import a sheet like your `Fichier_de_notes_M1NE...xlsx`:
        - Row 3 headers: Numéro, Nom, Prénom, Track, then assessment columns like "EE (0,4)", etc.
        - Row 2 indicates "session 1"/"session 2" regions.
        We import only assessment columns (not 'Note S1/S2').
        """
        path, _ = QFileDialog.getOpenFileName(
            self,
            "Importer des notes",
            str(Path.home() / "Documents"),
            "Excel (*.xlsx)",
        )
        if not path:
            return
        try:
            import openpyxl
        except Exception as exc:  # pragma: no cover
            QMessageBox.critical(self, "Dépendance manquante", f"openpyxl est requis.\n\n{exc}")
            return
        try:
            wb = openpyxl.load_workbook(path, data_only=True)
        except Exception as exc:
            QMessageBox.critical(self, "Import", str(exc))
            return

        try:
            sheet_names = wb.sheetnames
            # If only one sheet is relevant, choose active; else ask by simple heuristic (first non "Liste Etudiants")
            ws = wb.active
            if len(sheet_names) > 1:
                ws = wb[sheet_names[0]]
            # Try detect header row with "Numéro" and "Prénom"
            header_row_idx = None
            for i in range(1, min(20, ws.max_row) + 1):
                a = _norm_number(ws.cell(i, 1).value).lower()
                c = _norm_number(ws.cell(i, 3).value).lower()
                if "num" in a and ("prenom" in c or "prénom" in c):
                    header_row_idx = i
                    break
            if header_row_idx is None:
                QMessageBox.warning(self, "Import", "En-tête introuvable (attendu : Numéro/Nom/Prénom).")
                return

            # The row above header often contains session markers per column.
            session_row_idx = max(1, header_row_idx - 1)

            headers = [_norm_number(ws.cell(header_row_idx, c).value) for c in range(1, ws.max_column + 1)]
            sess_mark = [_norm_number(ws.cell(session_row_idx, c).value).lower() for c in range(1, ws.max_column + 1)]

            # Find student number column
            try:
                col_num = headers.index("Numéro") + 1
            except ValueError:
                col_num = 1

            # Build mapping from xlsx columns -> assessment_id
            # session from sess_mark cell: contains "session 1" or "session 2"
            by_key = {(m.session, m.kind, round(m.coefficient, 6)): m for m in self._col_maps}
            col_to_assessment: dict[int, int] = {}
            for idx, h in enumerate(headers, start=1):
                if not h or h.lower().startswith("note"):
                    continue
                parsed = _parse_assessment_header(h)
                if not parsed:
                    continue
                kind, coef = parsed
                sraw = sess_mark[idx - 1]
                session = 1 if "session 1" in sraw else (2 if "session 2" in sraw else 1)
                key = (int(session), kind, round(float(coef), 6))
                cm = by_key.get(key)
                if cm:
                    col_to_assessment[idx] = int(cm.assessment_id)

            if not col_to_assessment:
                QMessageBox.warning(
                    self,
                    "Import",
                    "Aucune colonne d'évaluation correspondante.\n"
                    "Astuce : générez d'abord les évaluations depuis les MCC, puis importez.",
                )
                return

            imported = 0
            skipped = 0
            errors: list[str] = []
            for r in range(header_row_idx + 1, ws.max_row + 1):
                stu_num = _norm_number(ws.cell(r, col_num).value)
                if not stu_num:
                    continue
                st = self.repo.get_student_by_number(stu_num)
                if not st:
                    skipped += 1
                    continue
                sid = int(st["id"])
                for cidx, aid in col_to_assessment.items():
                    v = ws.cell(r, cidx).value
                    if v is None or v == "" or v == -1:
                        continue
                    raw = str(v).strip() if not isinstance(v, (int, float)) else str(v)
                    grade, status, err = parse_grade_cell(raw)
                    if err:
                        errors.append(f"{stu_num} col {headers[cidx-1]!r}: {err}")
                        continue
                    try:
                        self.repo.upsert_grade(sid, int(aid), grade, status=status, comment="")
                        imported += 1
                    except Exception as exc:
                        errors.append(f"{stu_num} col {headers[cidx-1]!r}: {exc}")

            self.refresh()
            msg = f"Imported grade cells: {imported}\nSkipped students not found: {skipped}"
            if errors:
                msg += "\n\nErrors:\n" + "\n".join(errors[:12])
                if len(errors) > 12:
                    msg += f"\n… ({len(errors)-12} more)"
            msg_fr = (
                f"Cellules de notes importées : {imported}\n"
                f"Étudiants introuvables (ignorés) : {skipped}"
            )
            if errors:
                msg_fr += "\n\nErreurs:\n" + "\n".join(errors[:12])
                if len(errors) > 12:
                    msg_fr += f"\n… ({len(errors)-12} de plus)"
            QMessageBox.information(self, "Import terminé", msg_fr)
        finally:
            wb.close()

