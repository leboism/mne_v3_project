"""Import / modèle Excel pour les compositions du jury."""

from __future__ import annotations

import re
import unicodedata
from pathlib import Path
from typing import Any

JURY_MEMBER_HEADERS: tuple[str, ...] = (
    "last_name",
    "first_name",
    "title",
    "institution",
)

JURY_EXCEL_LABELS_FR: dict[str, str] = {
    "last_name": "Nom (obligatoire)",
    "first_name": "Prénom (obligatoire)",
    "title": "Qualité / titre (ex. Professeur, représentant professionnel)",
    "institution": "Institution",
}

JURY_EXCEL_EXAMPLE_ROWS: tuple[tuple[str, ...], ...] = (
    ("Oui", "Martin", "Sophie", "Professeure", "Université Paris-Saclay"),
    ("", "Durand", "Paul", "Représentant professionnel", "CEA"),
    ("", "Lefebvre", "Anne", "Étudiante", "Master MNE"),
)


def norm_header(value: Any) -> str:
    if value is None:
        return ""
    s = str(value).strip().lower()
    s = unicodedata.normalize("NFKD", s)
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    s = re.sub(r"[^a-z0-9]+", "_", s).strip("_")
    return s


def _idx_of(headers: list[str], *names: str) -> int | None:
    for n in names:
        n2 = norm_header(n)
        if n2 in headers:
            return headers.index(n2)
    return None


def is_president_marker(value: Any) -> bool:
    s = str(value or "").strip().lower()
    return s in ("1", "oui", "yes", "y", "x", "true", "vrai", "président", "president", "p", "o")


def split_jury_president_and_members(
    members: list[dict[str, Any]],
) -> tuple[dict[str, Any] | None, list[dict[str, Any]]]:
    """Retourne (président, autres membres). Si aucun désigné, le premier de la liste."""
    if not members:
        return None, []
    presidents = [m for m in members if int(m.get("is_president") or 0)]
    if presidents:
        president = presidents[0]
        others = [m for m in members if int(m.get("id") or 0) != int(president.get("id") or -1)]
        return president, others
    return members[0], members[1:]


def parse_jury_members_workbook(path: str | Path) -> tuple[list[dict[str, str]], list[str]]:
    """
    Lit un fichier Excel et retourne (membres, erreurs).
    Chaque membre : last_name, first_name, title, institution.
    """
    from openpyxl import load_workbook

    errors: list[str] = []
    wb = load_workbook(str(path), data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    header_row = next(rows_iter, None)
    if not header_row:
        return [], ["Fichier vide."]

    headers = [norm_header(h) for h in header_row]
    col_last = _idx_of(headers, "last_name", "nom", "surname", "lastname")
    col_first = _idx_of(headers, "first_name", "prenom", "firstname", "given_name")
    col_title = _idx_of(headers, "title", "qualite", "qualité", "fonction", "role")
    col_inst = _idx_of(headers, "institution", "etablissement", "établissement", "employeur")
    col_pres = _idx_of(headers, "president", "président", "is_president", "president_du_jury")

    missing = []
    if col_last is None:
        missing.append("last_name (nom)")
    if col_first is None:
        missing.append("first_name (prénom)")
    if missing:
        return [], [f"Colonnes obligatoires manquantes : {', '.join(missing)}."]

    def cell(row: tuple[Any, ...], idx: int | None) -> str:
        if idx is None or idx >= len(row):
            return ""
        v = row[idx]
        return "" if v is None else str(v).strip()

    members: list[dict[str, str]] = []
    for excel_row_idx, row in enumerate(rows_iter, start=2):
        last_name = cell(row, col_last)
        first_name = cell(row, col_first)
        if not last_name and not first_name:
            continue
        if not last_name or not first_name:
            errors.append(f"Ligne {excel_row_idx} : nom et prénom obligatoires.")
            continue
        members.append(
            {
                "last_name": last_name,
                "first_name": first_name,
                "title": cell(row, col_title),
                "institution": cell(row, col_inst),
                "is_president": 1 if is_president_marker(cell(row, col_pres)) else 0,
            }
        )

    if not members and not errors:
        errors.append("Aucun membre valide trouvé (vérifiez les lignes de données).")
    return members, errors


def write_jury_import_template(path: str | Path) -> None:
    """Modèle Excel : une ligne d'en-tête + exemples + onglet Instructions."""
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    ws = wb.active
    ws.title = "Jury"
    ws.append(["Président", *JURY_MEMBER_HEADERS])
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for row in JURY_EXCEL_EXAMPLE_ROWS:
        ws.append(list(row))
    ws.freeze_panes = "A2"

    ins = wb.create_sheet("Instructions")
    ins.append(["Colonne", "Description"])
    ins["A1"].font = Font(bold=True)
    ins["B1"].font = Font(bold=True)
    for key in JURY_MEMBER_HEADERS:
        ins.append([key, JURY_EXCEL_LABELS_FR.get(key, "")])
    ins.append(["president", "Oui / 1 / x pour désigner le président du jury (une seule ligne)"])
    ins.append([])
    ins.append(["Usage", "Une ligne = un membre du jury. Importez la liste complète d'un coup."])
    ins.append(["Alias FR", "nom, prénom, qualité, institution sont aussi reconnus."])

    wb.save(str(path))


def write_jury_roster_workbook(
    members: list[dict[str, str]],
    path: str | Path,
    *,
    title: str = "",
    academic_year: str = "",
) -> None:
    """Exporte une composition du jury (format réimportable)."""
    from datetime import date

    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    ws = wb.active
    ws.title = "Jury"
    headers_fr = ("Président", "Nom", "Prénom", "Qualité", "Institution")
    ws.append(list(headers_fr))
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for m in members:
        ws.append(
            [
                "Oui" if int(m.get("is_president") or 0) else "",
                str(m.get("last_name") or ""),
                str(m.get("first_name") or ""),
                str(m.get("title") or ""),
                str(m.get("institution") or ""),
            ]
        )
    ws.freeze_panes = "A2"

    if title or academic_year:
        ins = wb.create_sheet("Informations")
        ins["A1"].font = Font(bold=True)
        ins["B1"].font = Font(bold=True)
        ins.append(["Champ", "Valeur"])
        if title:
            ins.append(["Composition", title])
        if academic_year:
            ins.append(["Millésime", academic_year])
        ins.append(["Exporté le", date.today().isoformat()])
        ins.append([])
        ins.append(["Usage", "Fichier réimportable via « Importer composition (Excel) »."])

    wb.save(str(path))
