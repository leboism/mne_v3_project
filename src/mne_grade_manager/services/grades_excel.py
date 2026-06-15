"""Export / modèles Excel pour les notes (format compatible avec l'import MNE)."""

from __future__ import annotations

from pathlib import Path
from typing import TYPE_CHECKING, Any

from .grade_status import format_grade_display

if TYPE_CHECKING:
    from .repository import Repository

# Colonnes identité (lignes 3+ du fichier ; lignes 1–2 = session + en-têtes)
GRADES_IDENTITY_HEADERS = ("Numéro", "Nom", "Prénom", "Track")


def _coef_label(coef: float) -> str:
    """Format attendu par l'import : ``EE (0,4)``."""
    s = f"{float(coef):g}".replace(".", ",")
    return s


def assessment_header_label(kind: str, coefficient: float) -> str:
    return f"{str(kind).strip().upper()} ({_coef_label(coefficient)})"


def write_grades_workbook(
    repo: Repository,
    *,
    template_id: int,
    course_id: int,
    path: str | Path,
    students: list[dict[str, Any]],
    fill_grades: bool = False,
    include_instructions: bool = False,
) -> None:
    """
    Génère un Excel importable via « Saisie par matière → Importer ».

    - Ligne 1 : marqueurs « session 1 » / « session 2 » au-dessus des colonnes d'évaluations
    - Ligne 2 : Numéro, Nom, Prénom, Track, puis en-têtes du type ``EE (0,4)``
    - Ligne 3+ : une ligne par étudiant
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font

    course = repo.get_course(int(course_id)) or {}
    assessments = repo.list_assessments(int(course_id))
    if not assessments:
        raise ValueError(
            "Aucune évaluation pour cette UE. Générez les assessments depuis les MCC d'abord."
        )

    wb = Workbook()
    code = str(course.get("code") or "notes")[:31]
    ws = wb.active
    ws.title = code

    n_id = len(GRADES_IDENTITY_HEADERS)
    n_assess = len(assessments)
    n_cols = n_id + n_assess

    session_row = [""] * n_cols
    header_row = list(GRADES_IDENTITY_HEADERS)
    for a in assessments:
        session_row.append(f"session {int(a['session'])}")
        header_row.append(assessment_header_label(str(a["kind"]), float(a["coefficient"])))

    ws.append(session_row)
    ws.append(header_row)
    for c in range(1, n_cols + 1):
        ws.cell(2, c).font = Font(bold=True)

    grade_cache: dict[tuple[int, int], dict[str, Any]] = {}
    if fill_grades:
        for s in students:
            sid = int(s["id"])
            for g in repo.get_grades_for_student_course(sid, int(course_id)):
                grade_cache[(sid, int(g["assessment_id"]))] = g

    for s in students:
        sid = int(s["id"])
        row: list[Any] = [
            s.get("student_number", ""),
            s.get("last_name", ""),
            s.get("first_name", ""),
            s.get("track", ""),
        ]
        for a in assessments:
            if fill_grades:
                g = grade_cache.get((sid, int(a["id"])))
                if g:
                    row.append(
                        format_grade_display(
                            g.get("grade"),
                            g.get("status"),
                            assessment_session=int(a.get("session") or 1),
                        )
                    )
                else:
                    row.append("")
            else:
                row.append("")
        ws.append(row)

    ws.freeze_panes = "A3"

    if include_instructions:
        ins = wb.create_sheet("Instructions")
        ins.append(["Élément", "Description"])
        ins["A1"].font = Font(bold=True)
        ins["B1"].font = Font(bold=True)
        rows = [
            ("Ligne 1", "Indique session 1 ou session 2 pour chaque colonne d'évaluation."),
            ("Ligne 2", "En-têtes : ne pas modifier les colonnes EE (0,4), CC (0,3), etc."),
            ("Numéro", "Doit correspondre au student_number en base."),
            ("Notes", "Nombre /20, ou ABJ (absence justifiée), ou DEF (défaillant). Cellule vide = pas de note."),
            ("Import", "Notes → Saisie par matière → Importer un fichier de notes Excel."),
            ("UE", f"{course.get('code', '')} — {course.get('name', '')}"),
        ]
        for a, b in rows:
            ins.append([a, b])

    wb.save(str(path))


def write_grades_import_template(
    repo: Repository,
    *,
    template_id: int,
    course_id: int,
    path: str | Path,
    students: list[dict[str, Any]],
) -> None:
    """Modèle vide (étudiants listés, notes vides) + instructions."""
    write_grades_workbook(
        repo,
        template_id=int(template_id),
        course_id=int(course_id),
        path=path,
        students=students,
        fill_grades=False,
        include_instructions=True,
    )
