"""Export Excel « Fichier de notes » (format secrétariat M1NE)."""

from __future__ import annotations

from pathlib import Path
from typing import TYPE_CHECKING, Any

from .grade_status import (
    STATUS_ABJ,
    STATUS_DEF,
    STATUS_NEUT,
    STATUS_VAL,
    is_s2_reprise_assessment,
    normalize_grade_status,
)
from .grades_excel import GRADES_IDENTITY_HEADERS, assessment_header_label
from .timetable_legacy import course_public_code
from .dates import suggest_next_academic_year

if TYPE_CHECKING:
    from .repository import Repository

_NO_S2_NOTE = -1


def suggest_notes_workbook_filename(*, academic_year: str, level: str = "M1") -> str:
    """Ex. ``Fichier_de_notes_M1NE_2025-26.xlsx``."""
    ay = str(academic_year or "").strip()
    parts = ay.split("-")
    if len(parts) == 2 and len(parts[1]) == 4:
        short = f"{parts[0]}-{parts[1][-2:]}"
    else:
        short = ay.replace("/", "-")
    level_u = str(level or "M1").strip().upper()
    prefix = f"{level_u}NE" if level_u in {"M1", "M2"} else level_u
    return f"Fichier_de_notes_{prefix}_{short}.xlsx"


def _coef_label(coef: float) -> str:
    return f"{float(coef):g}".replace(".", ",")


def _jury_points_label(points: float) -> str:
    p = float(points)
    if abs(p) < 1e-12:
        return ""
    sign = "+" if p > 0 else ""
    return f"{sign}{_coef_label(p)}"


def grade_excel_value(
    grade: Any,
    status: Any,
    *,
    assessment_session: int | None = None,
    assessment_kind: str | None = None,
    assessment_name: str | None = None,
) -> Any:
    """Valeur cellule Excel pour une note d'épreuve."""
    st = normalize_grade_status(status)
    if (
        assessment_session == 2
        and grade is None
        and st == STATUS_DEF
        and is_s2_reprise_assessment(
            session=assessment_session,
            kind=assessment_kind,
            name=assessment_name,
        )
    ):
        return None
    if st in {STATUS_ABJ, STATUS_DEF, STATUS_NEUT, STATUS_VAL}:
        return st
    if grade is None:
        return None
    try:
        return float(grade)
    except (TypeError, ValueError):
        return str(grade)


def _course_sheet_name(course: dict[str, Any], *, academic_year: str) -> str:
    name = str(course.get("name") or "").lower()
    if "intern" in name:
        return "INTERNSHIP"
    pub = course_public_code(course, academic_year=academic_year)
    return (pub or str(course.get("code") or "UE"))[:31]


def _course_sort_key(course: dict[str, Any], *, academic_year: str) -> tuple[Any, ...]:
    pub = course_public_code(course, academic_year=academic_year).upper()
    if pub == "INTERNSHIP" or "INTERN" in pub:
        return (3, 0, pub)
    parts = pub.split("-")
    sem = parts[0] if parts else "Z"
    track = parts[1] if len(parts) > 1 else "Z"
    track_ord = {"C": 0, "P": 1, "X": 2}.get(track, 9)
    sem_ord = 0 if sem == "S1" else 1 if sem == "S2" else 2
    return (sem_ord, track_ord, pub)


def _course_title_cell(course: dict[str, Any]) -> str:
    code = str(course.get("code") or "").strip()
    name = str(course.get("name") or "").strip()
    if code and name:
        return f"{code} - {name}"
    return code or name


def _split_assessments(assessments: list[dict[str, Any]]) -> tuple[list[dict], list[dict]]:
    s1 = [a for a in assessments if int(a.get("session") or 1) == 1]
    s2 = [a for a in assessments if int(a.get("session") or 1) == 2]
    return s1, s2


def _grades_map(repo: Repository, student_id: int, course_id: int) -> dict[int, dict[str, Any]]:
    return {
        int(g["assessment_id"]): g
        for g in repo.get_grades_for_student_course(int(student_id), int(course_id))
    }


def _has_session2_activity(
    assessments_s2: list[dict[str, Any]],
    grades: dict[int, dict[str, Any]],
) -> bool:
    if not assessments_s2:
        return False
    for a in assessments_s2:
        g = grades.get(int(a["id"]), {})
        if g.get("grade") is not None:
            return True
        st = normalize_grade_status(g.get("status"))
        if st in {STATUS_ABJ, STATUS_DEF, STATUS_NEUT, STATUS_VAL}:
            return True
    return False


def _load_jury_course_cells(
    repo: Repository,
    *,
    template_ids: list[int],
    course_id: int,
) -> dict[int, Any]:
    if not template_ids:
        return {}
    placeholders = ",".join("?" * len(template_ids))
    rows = repo.db.query_all(
        f"""
        SELECT student_id, points, comment
        FROM jury_adjustments
        WHERE template_id IN ({placeholders})
          AND scope = 'course'
          AND course_id = ?
        """,
        (*[int(t) for t in template_ids], int(course_id)),
    )
    out: dict[int, Any] = {}
    for r in rows:
        sid = int(r["student_id"])
        pts = float(r["points"] or 0)
        comment = str(r["comment"] or "").strip()
        label = _jury_points_label(pts)
        if comment and not label:
            cell = comment
        elif comment and label:
            cell = comment if comment.startswith(("+", "-")) else f"{label} — {comment}"
        else:
            cell = label or (float(pts) if abs(pts) >= 1e-12 else None)
        if cell not in (None, ""):
            out[sid] = cell
    return out


def _student_included_in_level_export(
    student: dict[str, Any],
    *,
    academic_year: str,
    level: str,
) -> bool:
    """Exclut les étudiants déjà passés en M2 (millésime suivant) d'un export M1."""
    lv = str(level or "").strip().upper()
    st_level = str(student.get("level") or "").strip().upper()
    st_year = str(student.get("academic_year") or "").strip()
    ay = str(academic_year or "").strip()
    if lv == "M1" and st_level == "M2":
        next_ay = suggest_next_academic_year(ay)
        if next_ay and st_year == next_ay:
            return False
    return True


def _collect_export_context(
    repo: Repository,
    *,
    academic_year: str,
    level: str,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]], dict[str, int]]:
    templates = repo.list_templates_for_year_level(academic_year, level)
    if not templates:
        raise ValueError(f"Aucune maquette pour {level} — {academic_year}.")
    students_by_id: dict[int, dict[str, Any]] = {}
    for t in templates:
        for s in repo.list_students_for_template(int(t["id"])):
            if not _student_included_in_level_export(
                s, academic_year=academic_year, level=level
            ):
                continue
            students_by_id[int(s["id"])] = s
    students = sorted(
        students_by_id.values(),
        key=lambda s: (str(s.get("last_name") or ""), str(s.get("first_name") or "")),
    )
    templates_by_track = {
        str(t.get("track") or "").strip().upper(): int(t["id"]) for t in templates
    }
    courses_by_id: dict[int, dict[str, Any]] = {}
    for t in templates:
        ay = str(t.get("academic_year") or academic_year).strip()
        for c in repo.list_template_courses(int(t["id"])):
            cid = int(c["course_id"])
            if cid not in courses_by_id:
                courses_by_id[cid] = dict(c)
                courses_by_id[cid]["_export_ay"] = ay
    courses = sorted(
        courses_by_id.values(),
        key=lambda c: _course_sort_key(c, academic_year=str(c.get("_export_ay") or academic_year)),
    )
    return students, courses, templates_by_track


def _write_student_list_sheet(ws, students: list[dict[str, Any]]) -> None:
    ws.append([None, None, None, None, None])
    ws.append([None, None, None, None, None])
    ws.append([*GRADES_IDENTITY_HEADERS, None])
    for s in students:
        ws.append(
            [
                s.get("student_number", ""),
                s.get("last_name", ""),
                s.get("first_name", ""),
                s.get("track", ""),
                None,
            ]
        )


def _write_course_sheet(
    ws,
    repo: Repository,
    *,
    course: dict[str, Any],
    students: list[dict[str, Any]],
    academic_year: str,
    templates_by_track: dict[str, int],
    template_ids: list[int],
) -> None:
    from openpyxl.styles import Font

    course_id = int(course["course_id"])
    assessments = repo.list_assessments(course_id)
    if not assessments:
        return
    assess_s1, assess_s2 = _split_assessments(assessments)
    jury_cells = _load_jury_course_cells(
        repo, template_ids=template_ids, course_id=course_id
    )
    include_jury = bool(jury_cells)

    n_id = len(GRADES_IDENTITY_HEADERS)
    s1_start = n_id + 1
    note_s1_col = s1_start + len(assess_s1)
    s2_start = note_s1_col + 1
    note_s2_col = s2_start + len(assess_s2)
    jury_col = note_s2_col + 1 if include_jury else None
    n_cols = jury_col or note_s2_col

    title_row: list[Any] = [None] * n_cols
    title_row[n_id] = _course_title_cell(course)
    ws.append(title_row)

    session_row: list[Any] = [None] * n_cols
    if assess_s1:
        session_row[s1_start - 1] = "session 1"
    if assess_s2:
        session_row[s2_start - 1] = "session 2"
    ws.append(session_row)

    header_row: list[Any] = list(GRADES_IDENTITY_HEADERS)
    for a in assess_s1:
        header_row.append(
            assessment_header_label(str(a["kind"]), float(a["coefficient"]))
        )
    header_row.append("Note S1")
    for a in assess_s2:
        header_row.append(
            assessment_header_label(str(a["kind"]), float(a["coefficient"]))
        )
    header_row.append("Note S2")
    if include_jury:
        header_row.append(None)
    ws.append(header_row)
    for c in range(1, len(header_row) + 1):
        ws.cell(3, c).font = Font(bold=True)

    for s in students:
        sid = int(s["id"])
        row: list[Any] = [
            s.get("student_number", ""),
            s.get("last_name", ""),
            s.get("first_name", ""),
            s.get("track", ""),
        ]
        grades = _grades_map(repo, sid, course_id)
        for a in assess_s1:
            g = grades.get(int(a["id"]), {})
            row.append(
                grade_excel_value(
                    g.get("grade"),
                    g.get("status"),
                    assessment_session=1,
                )
            )
        row.append(repo.compute_course_average_s1(sid, course_id))
        has_s2 = _has_session2_activity(assess_s2, grades)
        for a in assess_s2:
            g = grades.get(int(a["id"]), {})
            row.append(
                grade_excel_value(
                    g.get("grade"),
                    g.get("status"),
                    assessment_session=2,
                    assessment_kind=str(a.get("kind") or ""),
                    assessment_name=str(a.get("name") or ""),
                )
                if has_s2
                else None
            )
        if has_s2:
            tid = templates_by_track.get(str(s.get("track") or "").strip().upper())
            row.append(
                repo.compute_course_average_s2(
                    sid, course_id, template_id=tid
                )
            )
        else:
            row.append(_NO_S2_NOTE)
        if include_jury:
            row.append(jury_cells.get(sid))
        ws.append(row)


def _summary_value(row_data: dict[str, Any] | None, col: dict[str, Any]) -> Any:
    if not row_data:
        return None
    kind = col.get("kind")
    if kind == "ue":
        return (row_data.get("courses") or {}).get(f"c:{int(col['course_id'])}")
    if kind == "block_avg":
        return (row_data.get("blocks") or {}).get(str(col["block_name"]))
    if kind == "year_avg":
        return row_data.get("global_with_jury")
    return None


def _student_summary_lookup(
    repo: Repository,
    *,
    student: dict[str, Any],
    templates_by_track: dict[str, int],
    view_session: str,
) -> dict[str, Any] | None:
    tr = str(student.get("track") or "").strip().upper()
    tid = templates_by_track.get(tr)
    if tid is None:
        return None
    summaries = repo.get_student_result_summary(
        int(tid), view_session=view_session, include_all_students=True
    )
    return next(
        (r for r in summaries if int(r["student_id"]) == int(student["id"])),
        None,
    )


def _build_session_columns(
    repo: Repository,
    templates: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    """Colonnes UE + moyennes de bloc pour la feuille de synthèse."""
    columns: list[dict[str, Any]] = []
    seen_blocks: set[str] = set()
    seen_courses: set[int] = set()

    for tpl in templates:
        tid = int(tpl["id"])
        for bk, clist in repo.list_template_blocks_with_courses(tid):
            for c in clist:
                if int(c.get("optional") or 0):
                    continue
                cid = int(c["course_id"])
                if cid in seen_courses:
                    continue
                seen_courses.add(cid)
                columns.append({"kind": "ue", "course_id": cid, "course": c})
            if bk not in seen_blocks:
                seen_blocks.add(bk)
                columns.append({"kind": "block_avg", "block_name": bk})
                columns.append({"kind": "block_result", "block_name": bk})
    columns.append({"kind": "year_avg"})
    columns.append({"kind": "year_result"})
    return columns


def _write_session_summary_sheet(
    ws,
    repo: Repository,
    *,
    students: list[dict[str, Any]],
    templates: list[dict[str, Any]],
    templates_by_track: dict[str, int],
    academic_year: str,
    view_session: str,
    title: str,
) -> None:
    from openpyxl.styles import Font

    columns = _build_session_columns(repo, templates)
    n_id = 4
    n_cols = n_id + len(columns)

    row1: list[Any] = [None] * n_cols
    row1[n_id] = title
    ws.append(row1)

    row2: list[Any] = [None] * n_cols
    row3: list[Any] = [None] * n_cols
    row4: list[Any] = [None] * n_cols
    row4[0] = "Numéro"
    row4[1] = "Nom"
    row4[2] = "Prénom"
    row4[3] = "Track"

    col_idx = n_id
    for col in columns:
        col_idx += 1
        if col["kind"] == "ue":
            c = col["course"]
            row3[col_idx - 1] = _course_title_cell(c)
            row4[col_idx - 1] = float(c.get("ects") or c.get("global_coefficient") or 0) or None
        elif col["kind"] == "block_avg":
            row3[col_idx - 1] = "moyenne"
        elif col["kind"] == "block_result":
            row3[col_idx - 1] = "resultat"
        elif col["kind"] == "year_avg":
            row3[col_idx - 1] = "moyenne"
        elif col["kind"] == "year_result":
            row3[col_idx - 1] = "resultat"

    ws.append(row2)
    ws.append(row3)
    ws.append(row4)
    for c in range(1, 5):
        ws.cell(4, c).font = Font(bold=True)

    for s in students:
        tr = str(s.get("track") or "").strip().upper()
        row_data = _student_summary_lookup(
            repo,
            student=s,
            templates_by_track=templates_by_track,
            view_session=view_session,
        )
        row: list[Any] = [
            s.get("student_number", ""),
            s.get("last_name", ""),
            s.get("first_name", ""),
            tr,
        ]
        for col in columns:
            if col["kind"] in {"block_result", "year_result"}:
                row.append(None)
            else:
                row.append(_summary_value(row_data, col))
        ws.append(row)


def write_results_notes_workbook(
    repo: Repository,
    *,
    academic_year: str,
    level: str,
    path: str | Path,
) -> None:
    """
    Classeur multi-feuilles type « Fichier_de_notes_M1NE » :
    Liste Etudiants, une feuille par UE, synthèses session 1 / session 2.
    """
    from openpyxl import Workbook

    students, courses, templates_by_track = _collect_export_context(
        repo, academic_year=academic_year, level=level
    )
    templates = repo.list_templates_for_year_level(academic_year, level)
    template_ids = [int(t["id"]) for t in templates]
    ay = str(academic_year).strip()

    wb = Workbook()
    wb.remove(wb.active)

    ws_students = wb.create_sheet("Liste Etudiants", 0)
    _write_student_list_sheet(ws_students, students)

    for course in courses:
        sheet_name = _course_sheet_name(course, academic_year=ay)
        if sheet_name in wb.sheetnames:
            base = sheet_name[:28]
            n = 2
            while f"{base}_{n}" in wb.sheetnames:
                n += 1
            sheet_name = f"{base}_{n}"
        ws = wb.create_sheet(sheet_name)
        _write_course_sheet(
            ws,
            repo,
            course=course,
            students=students,
            academic_year=str(course.get("_export_ay") or ay),
            templates_by_track=templates_by_track,
            template_ids=template_ids,
        )

    ws_s1 = wb.create_sheet("session 1")
    _write_session_summary_sheet(
        ws_s1,
        repo,
        students=students,
        templates=templates,
        templates_by_track=templates_by_track,
        academic_year=ay,
        view_session="s1",
        title="session 1",
    )
    ws_s2 = wb.create_sheet("session 2")
    _write_session_summary_sheet(
        ws_s2,
        repo,
        students=students,
        templates=templates,
        templates_by_track=templates_by_track,
        academic_year=ay,
        view_session="mixed",
        title="session 2",
    )

    wb.save(str(path))
