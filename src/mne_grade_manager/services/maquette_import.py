"""Parse Excel « maquette » (ex. maquette 2024.xlsx) — structure UPSay / MNE."""

from __future__ import annotations

import re
import unicodedata
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable

# Noms d’onglets du fichier maquette → code parcours (templates.track / students.track)
# M1 : P et C ; M2 : NPD, NPO, DWM, NFC, NRPE
MAQUETTE_SHEET_TO_TRACK: dict[str, str] = {
    "M1P": "P",
    "M1C": "C",
    "M2 NPO": "NPO",
    "M2 NDWM": "DWM",
    "M2 NFC": "NFC",
    "M2 NRPE": "NRPE",
    "M2 NPD": "NPD",
}


def _cell_str(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _cell_float(value: Any) -> float:
    if value is None or value == "":
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).strip()
    # Ex: "3 ECTS", "18 ECTS", "27,5", "10"
    m = re.search(r"[-+]?\d+(?:[.,]\d+)?", s)
    if not m:
        return 0.0
    try:
        return float(m.group(0).replace(",", "."))
    except ValueError:
        return 0.0


def find_maquette_header_row(rows: list[tuple[Any, ...]]) -> int | None:
    """Index 0-based de la ligne d’en-tête (colonne A = « Code », B contient « Enseignements »)."""
    for i, row in enumerate(rows):
        if not row:
            continue
        a = _cell_str(row[0]).lower()
        b = _cell_str(row[1]).lower()
        if a == "code" and "enseignement" in b:
            return i
    return None


def parse_maquette_rows(sheet_rows: Iterable[tuple[Any, ...]]) -> list[dict[str, Any]]:
    """
    Extrait les lignes « cours » après l’en-tête maquette.
    Retourne des dicts prêts pour insert/update (clés alignées sur la table courses).
    """
    from ..core.mne_modules import normalize_mne_module_code, validate_mne_module_code

    rows = list(sheet_rows)
    hi = find_maquette_header_row(rows)
    if hi is None:
        return []
    data_start = hi + 2
    out: list[dict[str, Any]] = []
    current_block = ""
    display_order = 1

    header = rows[hi] if hi < len(rows) else ()
    subheader = rows[hi + 1] if (hi + 1) < len(rows) else ()

    def _norm(v: Any) -> str:
        return _cell_str(v).lower()

    def _header_label(idx: int) -> str:
        top = _norm(header[idx]) if idx < len(header) else ""
        sub = _norm(subheader[idx]) if idx < len(subheader) else ""
        return sub or top

    def _find_col(predicate) -> int | None:
        for i in range(max(len(header), len(subheader))):
            lab = _header_label(i)
            if lab and predicate(lab):
                return i
        return None

    code_i = _find_col(lambda s: s == "code") or 0
    name_i = _find_col(lambda s: "enseignement" in s) or 1
    ue_i = _find_col(lambda s: s == "code ue" or (s.startswith("code") and "ue" in s))
    nomen_i = _find_col(lambda s: "nomenclature" in s)
    sem_i = _find_col(lambda s: "semestre" in s)
    total_i = _find_col(lambda s: s == "total")
    ead_i = _find_col(lambda s: s == "ead")
    ects_i = _find_col(lambda s: "ects" in s)
    mcc_i = _find_col(lambda s: "contrôle" in s or "controle" in s)

    cm_i = _find_col(lambda s: "magistral" in s)
    td_i = _find_col(lambda s: "dirig" in s)
    tp_i = _find_col(lambda s: "pratique" in s)
    proj_i = _find_col(lambda s: s == "projet")
    pt_i = _find_col(lambda s: "tutor" in s)
    aa_i = _find_col(lambda s: "autonome" in s)

    for row in rows[data_start:]:
        if not row:
            continue
        code = _cell_str(row[code_i] if code_i < len(row) else "")
        name = _cell_str(row[name_i] if name_i < len(row) else "").replace("\n", " ")
        if not code or not name:
            continue
        # Lignes "bloc" / groupement: code BCxxxxxx (ne sont pas des UE)
        if code.strip().upper().startswith("BC"):
            current_block = name
            continue
        # UE conteneur Apogée (programme / parcours) — pas un module enseigné
        if code.strip().upper().startswith("CU"):
            continue

        semester = _cell_str(row[sem_i]) if (sem_i is not None and sem_i < len(row)) else ""
        cm = _cell_float(row[cm_i]) if (cm_i is not None and cm_i < len(row)) else 0.0
        td = _cell_float(row[td_i]) if (td_i is not None and td_i < len(row)) else 0.0
        tp = _cell_float(row[tp_i]) if (tp_i is not None and tp_i < len(row)) else 0.0
        proj = _cell_float(row[proj_i]) if (proj_i is not None and proj_i < len(row)) else 0.0
        pt = _cell_float(row[pt_i]) if (pt_i is not None and pt_i < len(row)) else 0.0
        aa = _cell_float(row[aa_i]) if (aa_i is not None and aa_i < len(row)) else 0.0
        total = _cell_float(row[total_i]) if (total_i is not None and total_i < len(row)) else 0.0
        ead = _cell_str(row[ead_i]) if (ead_i is not None and ead_i < len(row)) else ""
        ects = _cell_float(row[ects_i]) if (ects_i is not None and ects_i < len(row)) else 0.0
        mcc = _cell_str(row[mcc_i]) if (mcc_i is not None and mcc_i < len(row)) else ""

        mne_module_code = ""
        for idx in (nomen_i, ue_i):
            if idx is None or idx >= len(row):
                continue
            candidate = normalize_mne_module_code(_cell_str(row[idx]))
            if candidate and validate_mne_module_code(candidate):
                mne_module_code = candidate
                break
        if not mne_module_code and validate_mne_module_code(code):
            mne_module_code = normalize_mne_module_code(code)

        out.append(
            {
                "code": code,
                "name": name,
                "semester": semester,
                "hours_cm": cm,
                "hours_td": td,
                "hours_tp": tp,
                "hours_project": proj,
                "hours_pt": pt,
                "hours_aa": aa,
                "hours_total": total,
                "ects": ects,
                "ead_flag": ead,
                "code_other": "",
                "mcc_text": mcc,
                "mne_module_code": mne_module_code,
                # Placement (maquette): bloc = groupement d’UE, ordre d’affichage dans la feuille.
                "block_name": current_block,
                "display_order": display_order,
            }
        )
        display_order += 1
    return out


def enrich_maquette_rows_mne_codes(
    rows: list[dict[str, Any]],
    *,
    level: str,
    track: str,
) -> list[dict[str, Any]]:
    """Renseigne le Code UE (``mne_module_code``, ex. M1B1-C-THER) pour chaque ligne."""
    from ..core.mne_modules import (
        is_legacy_semester_ue_code,
        match_mne_module_code,
        normalize_mne_module_code,
        validate_mne_module_code,
    )

    out: list[dict[str, Any]] = []
    for row in rows:
        enriched = dict(row)
        code = str(enriched.get("code") or "").strip()
        if not str(enriched.get("mne_module_code") or "").strip():
            if validate_mne_module_code(code):
                enriched["mne_module_code"] = normalize_mne_module_code(code)
            else:
                mne = match_mne_module_code(
                    str(enriched.get("name") or ""),
                    block_name=str(enriched.get("block_name") or ""),
                    level=level,
                    track=track,
                )
                if mne and validate_mne_module_code(mne):
                    enriched["mne_module_code"] = mne
        other = str(enriched.get("code_other") or "").strip()
        if other and is_legacy_semester_ue_code(other):
            enriched["code_other"] = ""
        out.append(enriched)
    return out


@dataclass
class MaquetteParseResult:
    sheet_names: list[str]
    track_by_sheet: dict[str, str]
    rows: list[dict[str, Any]]
    sheet_title: str


def load_maquette_sheet(path: Path | str, sheet_name: str) -> MaquetteParseResult:
    from openpyxl import load_workbook

    p = Path(path)
    wb = load_workbook(p, read_only=True, data_only=True)
    try:
        sheet_names = list(wb.sheetnames)
        if sheet_name not in sheet_names:
            raise ValueError(f"Onglet inconnu: {sheet_name!r}. Disponibles: {sheet_names}")
        ws = wb[sheet_name]
        raw = list(ws.iter_rows(values_only=True))
    finally:
        wb.close()
    parsed = parse_maquette_rows(raw)
    return MaquetteParseResult(
        sheet_names=sheet_names,
        track_by_sheet=dict(MAQUETTE_SHEET_TO_TRACK),
        rows=parsed,
        sheet_title=sheet_name,
    )


def list_maquette_sheets(path: Path | str) -> list[str]:
    from openpyxl import load_workbook

    p = Path(path)
    wb = load_workbook(p, read_only=True, data_only=True)
    try:
        return list(wb.sheetnames)
    finally:
        wb.close()


M1_TRACKS: tuple[str, ...] = ("P", "C")
M2_TRACKS: tuple[str, ...] = ("NPD", "NPO", "DWM", "NFC", "NRPE")

# Alias présents dans les blocs OF (ex. « NDWM » dans le fichier officiel).
_M2_TOKEN_TO_TRACK: tuple[tuple[str, str], ...] = (
    ("NDWM", "DWM"),
    ("NRPE", "NRPE"),
    ("NPD", "NPD"),
    ("NPO", "NPO"),
    ("NFC", "NFC"),
    ("DWM", "DWM"),
)


def _ascii_upper(value: str) -> str:
    s = unicodedata.normalize("NFKD", value or "")
    return "".join(ch for ch in s if not unicodedata.combining(ch)).upper()


def extract_academic_year_from_path(path: Path | str) -> str:
    """Extrait 2026-2027 depuis le nom de fichier (prioritaire) ou le chemin."""
    p = Path(path)
    pattern = re.compile(r"(20\d{2})\s*[-–]\s*(20\d{2})")

    def _year_span(text: str) -> str:
        m = pattern.search(text)
        if not m:
            return ""
        y1, y2 = int(m.group(1)), int(m.group(2))
        if y2 == y1 + 1:
            return f"{y1}-{y2}"
        return ""

    # Nom de fichier d'abord (évite « Maquette 2026-2031 » dans le dossier parent).
    year = _year_span(p.stem)
    if year:
        return year
    m = pattern.search(str(p))
    if m:
        y1, y2 = int(m.group(1)), int(m.group(2))
        if y2 == y1 + 1:
            return f"{y1}-{y2}"
    return ""


def is_consolidated_of_sheet(sheet_name: str, file_path: Path | str | None = None) -> bool:
    """Fichiers OF type PR1162 / PR1163 (un onglet, tous parcours)."""
    u = _ascii_upper(sheet_name)
    if u.startswith("PR116") or u.startswith("OF_PR"):
        return True
    if file_path is not None:
        stem = _ascii_upper(Path(file_path).stem)
        if "OF_PR1162" in stem or "OF_PR1163" in stem:
            return True
        if "PR1162" in stem or "PR1163" in stem:
            return True
    return False


def infer_of_level(sheet_name: str, file_path: Path | str | None = None) -> str:
    blob = _ascii_upper(f"{sheet_name} {Path(file_path).name if file_path else ''}")
    if "1162" in blob or re.search(r"\bM1\b", blob):
        return "M1"
    if "1163" in blob or re.search(r"\bM2\b", blob):
        return "M2"
    return ""


def _tracks_from_parenthetical(name: str) -> set[str] | None:
    """Extrait NPD, NPO, … depuis un libellé du type « … (NPD-NPO-NDWM) »."""
    m = re.search(r"\(([^)]+)\)", name or "")
    if not m:
        return None
    blob = _ascii_upper(m.group(1))
    found: set[str] = set()
    for alias, track in _M2_TOKEN_TO_TRACK:
        if re.search(rf"(^|[^A-Z]){re.escape(alias)}([^A-Z]|$)", blob):
            found.add(track)
    return found or None


def infer_row_tracks(row: dict[str, Any], level: str) -> set[str]:
    """Parcours concernés par une ligne (tronc commun → tous les parcours du niveau)."""
    lv = (level or "").strip().upper()
    block = _ascii_upper(_cell_str(row.get("block_name")))
    name = _cell_str(row.get("name"))
    name_up = _ascii_upper(name)
    text = f"{block} {name_up}"

    if lv == "M1":
        if "SPECIALITE CHIMIE" in block or "SPECIALITE CHIMIE" in text:
            return {"C"}
        if "SPECIALITE PHYSIQUE" in block or "SPECIALITE PHYSIQUE" in text:
            return {"P"}
        return set(M1_TRACKS)

    if lv == "M2":
        m = re.search(r"BLOC\s+\d+\s+([A-Z]+)", block)
        if m:
            tok = m.group(1)
            for alias, track in _M2_TOKEN_TO_TRACK:
                if tok == alias:
                    return {track}
        # Priorité au libellé du cours (ex. « Risk Mgmt (NPD-NPO-NDWM-NFC) »),
        # pas à l’en-tête de bloc qui peut lister tous les parcours.
        restricted = _tracks_from_parenthetical(name)
        if restricted is not None:
            return restricted
        found: set[str] = set()
        for alias, track in _M2_TOKEN_TO_TRACK:
            if re.search(rf"(^|[^A-Z]){re.escape(alias)}([^A-Z]|$)", name_up):
                found.add(track)
        if found:
            return found
        if "INTERNSHIP" in name_up or "STAGE" in block:
            return set(M2_TRACKS)
        return set(M2_TRACKS)

    return set()


def split_rows_by_track(rows: list[dict[str, Any]], level: str) -> dict[str, list[dict[str, Any]]]:
    """Répartit les UE d'une OF consolidée par parcours (déduplication par code)."""
    lv = (level or "").strip().upper()
    tracks = list(M1_TRACKS if lv == "M1" else M2_TRACKS if lv == "M2" else ())
    out: dict[str, list[dict[str, Any]]] = {t: [] for t in tracks}
    seen: dict[str, set[str]] = {t: set() for t in tracks}

    for row in rows:
        code = _cell_str(row.get("code"))
        if not code:
            continue
        for track in infer_row_tracks(row, lv):
            if track not in out or code in seen[track]:
                continue
            seen[track].add(code)
            out[track].append(dict(row))

    for track in tracks:
        for i, row in enumerate(out[track], start=1):
            row["display_order"] = i
    return {t: out[t] for t in tracks if out[t]}


@dataclass
class ConsolidatedTrackPlan:
    level: str
    track: str
    name: str
    sheet: str
    rows: list[dict[str, Any]]

    @property
    def row_count(self) -> int:
        return len(self.rows)


def plan_consolidated_of_import(
    path: Path | str,
    sheet_name: str,
    *,
    academic_year: str = "",
) -> list[ConsolidatedTrackPlan]:
    """Prépare un plan d'import : 1 maquette par parcours à partir d'une OF consolidée."""
    result = load_maquette_sheet(path, sheet_name)
    year = (academic_year or "").strip() or extract_academic_year_from_path(path)
    level = infer_of_level(sheet_name, path)
    if not level:
        raise ValueError(
            "Impossible de déterminer le niveau (M1/M2). Indiquez l'année ou utilisez un fichier PR1162/PR1163."
        )
    split = split_rows_by_track(result.rows, level)
    plans: list[ConsolidatedTrackPlan] = []
    for track in (M1_TRACKS if level == "M1" else M2_TRACKS):
        rows = split.get(track) or []
        if not rows:
            continue
        rows = enrich_maquette_rows_mne_codes(rows, level=level, track=track)
        name = f"{year} — {level} {track}".strip(" —") if year else f"{level} {track}"
        plans.append(
            ConsolidatedTrackPlan(
                level=level,
                track=track,
                name=name,
                sheet=sheet_name,
                rows=rows,
            )
        )
    return plans


def detect_maquette_import_mode(path: Path | str, sheet_names: list[str] | None = None) -> str:
    """
    of_consolidated | multi | single
    """
    names = sheet_names or list_maquette_sheets(path)
    if len(names) == 1 and is_consolidated_of_sheet(names[0], path):
        return "of_consolidated"
    if any(n in MAQUETTE_SHEET_TO_TRACK for n in names):
        return "multi"
    return "single"
