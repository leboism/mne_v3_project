"""Import Excel emploi du temps (format secrétariat MNE : Timetable-M1-2024-25- v1.xlsx)."""

from __future__ import annotations

import re
import warnings
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path
from typing import Any

from .timetable_legacy import map_legacy_timetable_code, normalize_legacy_code

_WEEKDAYS = ("Monday", "Tuesday", "Wednesday", "Thursday", "Friday")
_LEGACY_IN_CELL_RE = re.compile(r"(S[1-4]-[CPX]-[A-Z0-9]+)", re.IGNORECASE)
_MNE_IN_CELL_RE = re.compile(
    r"(M[12](?:B[1-4]|-B[1-4]|-B\d)-[A-Z0-9-]+)",
    re.IGNORECASE,
)
_GENERIC_SUPERVISOR_SHEET_RE = re.compile(
    r"^Supervisors?\s+and\s+lectur(?:e|er)s?$",
    re.IGNORECASE,
)

_GRID_SHEET_RE = re.compile(
    r"^(M[12])\s+(Physics|Chemistry)\s+(S[12])$",
    re.IGNORECASE,
)
_SUPERVISOR_SHEET_RE = re.compile(
    r"^Supervisors?\s+and\s+lectur(?:e|er)s?\s+(S[12])$",
    re.IGNORECASE,
)
_M2_COMMON_TERM_SHEET_RE = re.compile(
    r"^(20\d{2}\s*[-–]\s*20\d{2})_(1st|2nd)_semestre$",
    re.IGNORECASE,
)
_M2_COURSES_SHEET_RE = re.compile(r"^Courses\s+(20\d{2}\s*[-–]\s*20\d{2})$", re.IGNORECASE)

_SLOT_HOURS: dict[str, float] = {
    "9:00-12:15": 3.25,
    "1:15-4:30": 3.25,
    "13:15-16:30": 3.25,
}


@dataclass
class TimetableReferenceRow:
    period: str
    block_label: str
    course_title: str
    legacy_code: str
    mne_module_code: str
    supervisors: str
    hours_expected: float
    ects: float


@dataclass
class TimetableSlotRow:
    level: str
    track: str
    period: str
    week_label: str
    week_number: int
    week_start_date: str
    day_of_week: str
    time_slot: str
    raw_text: str
    legacy_code: str
    mne_module_code: str
    teacher_initials: str
    room: str
    slot_kind: str
    fill_color: str = ""


@dataclass
class TimetableImportResult:
    academic_year: str
    level: str
    source_filename: str
    reference_courses: list[TimetableReferenceRow] = field(default_factory=list)
    slots: list[TimetableSlotRow] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)


def _cell_str(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return str(value).strip()


def _cell_float(value: Any) -> float:
    if value is None or value == "":
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    m = re.search(r"[-+]?\d+(?:[.,]\d+)?", str(value))
    if not m:
        return 0.0
    try:
        return float(m.group(0).replace(",", "."))
    except ValueError:
        return 0.0


def _extract_year_from_text(text: str) -> str:
    """Extrait un millésime depuis un libellé (fichier, onglet, titre…)."""
    blob = _cell_str(text)
    if not blob:
        return ""

    m = re.search(r"(20\d{2})\s*[-–]\s*(20\d{2})", blob)
    if m:
        y1, y2 = int(m.group(1)), int(m.group(2))
        if y2 == y1 + 1:
            return f"{y1}-{y2}"

    m = re.search(r"(20\d{2})\s*[-–]\s*(\d{2})(?:\D|$)", blob)
    if m:
        y1, y2s = int(m.group(1)), int(m.group(2))
        y2 = y1 // 100 * 100 + int(y2s)
        if y2 < y1:
            y2 += 100
        if y2 == y1 + 1:
            return f"{y1}-{y2}"

    # Abrégé secrétariat : « -26-27 », « 24-25 », etc.
    m = re.search(r"(?:^|[^\d])(\d{2})\s*[-–]\s*(\d{2})(?:[^\d]|$)", blob)
    if m:
        y1, y2 = int(m.group(1)), int(m.group(2))
        if y2 == y1 + 1:
            return f"20{y1}-20{y2}"

    return ""


def extract_academic_year_from_timetable(
    path: Path | str,
    workbook_title: str = "",
    *,
    sheet_names: list[str] | None = None,
) -> str:
    """Extrait 2026-2027 depuis le nom de fichier, les onglets ou le titre Excel."""
    from .maquette_import import extract_academic_year_from_path

    year = extract_academic_year_from_path(path)
    if year:
        return year

    p = Path(path)
    year = _extract_year_from_text(p.stem)
    if year:
        return year

    year = _extract_year_from_text(_cell_str(workbook_title))
    if year:
        return year

    for name in sheet_names or []:
        year = _extract_year_from_text(name)
        if year:
            return year

    return _extract_year_from_text(str(p))


def normalize_time_slot(raw: str) -> str:
    s = re.sub(r"\s+", "", (raw or "").strip().lower())
    s = s.replace(".", ":")
    if s.startswith("13:") or s.startswith("1:15"):
        return "1:15-4:30"
    if s.startswith("9:"):
        return "9:00-12:15"
    if "9:00" in s and "12:15" in s:
        return "9:00-12:15"
    if ("1:15" in s or "13:15" in s) and ("4:30" in s or "16:30" in s):
        return "1:15-4:30"
    return ""


def _parse_grid_sheet_meta(sheet_name: str) -> tuple[str, str, str] | None:
    m = _GRID_SHEET_RE.match(sheet_name.strip())
    if not m:
        return None
    level = m.group(1).upper()
    track = "P" if m.group(2).lower() == "physics" else "C"
    period = m.group(3).upper()
    return level, track, period


def _parse_m2_common_grid_sheet_meta(sheet_name: str) -> tuple[str, str, str] | None:
    """
    Format secrétariat alternatif (M2 common track) :
    onglets du type « 2026-2027_1st_semestre » et « 2026-2027_2nd_semestre ».
    """
    m = _M2_COMMON_TERM_SHEET_RE.match(sheet_name.strip())
    if not m:
        return None
    term = m.group(2).lower()
    period = "S1" if term.startswith("1") else "S2"
    return "M2", "X", period


def _parse_supervisor_sheet_meta(sheet_name: str) -> str | None:
    m = _SUPERVISOR_SHEET_RE.match(sheet_name.strip())
    return m.group(1).upper() if m else None


def _is_m2_courses_sheet(sheet_name: str) -> bool:
    return bool(_M2_COURSES_SHEET_RE.match(sheet_name.strip()))


def parse_m2_courses_sheet(rows: list[tuple[Any, ...]]) -> list[TimetableReferenceRow]:
    """
    Feuille « Courses 2026-2027 » (M2 common track) :
    colonnes typiques : COURSE | CODE | SUPERVISORS AND LECTURER(S) | HOURS | ...
    """
    header_idx = None
    for i, row in enumerate(rows):
        if not row:
            continue
        if _cell_str(row[0]).strip().upper() == "COURSE" and _cell_str(row[1]).strip().upper() == "CODE":
            header_idx = i
            break
    if header_idx is None:
        return []

    out: list[TimetableReferenceRow] = []
    for row in rows[header_idx + 1 :]:
        if not row:
            continue
        title = _cell_str(row[0] if len(row) > 0 else "")
        code = _cell_str(row[1] if len(row) > 1 else "")
        sup = _cell_str(row[2] if len(row) > 2 else "")
        hrs = _cell_float(row[3] if len(row) > 3 else 0)
        if not (title or code):
            continue
        # Ici le "CODE" est déjà un code MNE (M2B1-..., etc.). Pas de legacy.
        out.append(
            TimetableReferenceRow(
                period="",
                block_label="",
                course_title=title,
                legacy_code="",
                mne_module_code=code,
                supervisors=sup,
                hours_expected=hrs,
                ects=0.0,
            )
        )
    return out


def _classify_cell(raw: str, legacy_code: str) -> str:
    blob = (raw or "").upper()
    if not blob.strip():
        return "empty"
    if legacy_code:
        if "EXAM" in blob:
            return "exam"
        return "course"
    if any(k in blob for k in ("HOLIDAY", "BANK HOLIDAY", "VACANCE")):
        return "holiday"
    if "EXAM" in blob:
        return "exam"
    if any(k in blob for k in ("WELCOME", "WORKSHOP", "FORUM", "PRESENTATION", "RENTREE")):
        return "event"
    return "other"


def _infer_period_from_code(code: str) -> str:
    c = (code or "").strip().upper()
    m = re.match(r"^S([12])-", c)
    if m:
        return f"S{m.group(1)}"
    return ""


def _parse_cell_content(raw: str) -> tuple[str, str, str]:
    text = (raw or "").strip()
    if not text:
        return "", "", ""
    legacy = ""
    mne_direct = ""
    m = _LEGACY_IN_CELL_RE.search(text)
    if m:
        legacy = normalize_legacy_code(m.group(1))
    else:
        mm = _MNE_IN_CELL_RE.search(text)
        if mm:
            mne_direct = mm.group(1).strip().upper()
    lines = [ln.strip() for ln in text.splitlines() if ln.strip()]
    teacher = ""
    room = ""
    if legacy:
        for ln in lines:
            if legacy.upper() in ln.upper().replace(" ", ""):
                rest = re.sub(re.escape(legacy), "", ln, flags=re.IGNORECASE).strip()
                if rest:
                    teacher = rest
                break
        if not teacher and len(lines) >= 2:
            second = lines[1]
            if not second.upper().startswith("S"):
                teacher = second
    if "HBAR" in text.upper() or "ROOM" in text.upper():
        room_m = re.search(r"(HBAR[^\\n]+|Room[s]?[^\\n]+)", text, re.IGNORECASE)
        if room_m:
            room = room_m.group(1).strip()
    mne = map_legacy_timetable_code(legacy) if legacy else mne_direct
    if not legacy and mne_direct:
        teacher = teacher or _teacher_from_mne_cell(text, mne_direct)
    return legacy, mne, teacher


def _teacher_from_mne_cell(text: str, mne_code: str) -> str:
    first = (text or "").splitlines()[0].strip()
    rest = re.sub(re.escape(mne_code), "", first, flags=re.IGNORECASE).strip(" -–—\t")
    return rest


def _is_generic_supervisor_sheet(sheet_name: str) -> bool:
    return bool(_GENERIC_SUPERVISOR_SHEET_RE.match(sheet_name.strip()))


def parse_supervisor_sheet(
    rows: list[tuple[Any, ...]], period: str = ""
) -> list[TimetableReferenceRow]:
    """Feuille référentiel (codes EdT, heures prévues, intervenants)."""
    header_idx = None
    for i, row in enumerate(rows):
        if not row:
            continue
        cells = [_cell_str(c).upper() for c in row]
        if "COURSE" in cells and ("HOURS" in cells or "HOUR" in cells):
            header_idx = i
            break
        if any(c == "CODE" for c in cells):
            header_idx = i
            break
    if header_idx is None:
        return []

    out: list[TimetableReferenceRow] = []
    block_label = ""
    for row in rows[header_idx + 1 :]:
        if not row:
            continue
        block_cell = _cell_str(row[0] if len(row) > 0 else "")
        if block_cell.upper().startswith("BLOCK"):
            block_label = block_cell.replace("\n", " ").strip()
            continue
        title = _cell_str(row[3] if len(row) > 3 else "")
        legacy = normalize_legacy_code(_cell_str(row[4] if len(row) > 4 else ""))
        if not legacy and not title:
            continue
        if not legacy:
            continue
        row_period = period or _infer_period_from_code(legacy)
        mne = map_legacy_timetable_code(legacy)
        out.append(
            TimetableReferenceRow(
                period=row_period,
                block_label=block_label,
                course_title=title,
                legacy_code=legacy,
                mne_module_code=mne,
                supervisors=_cell_str(row[5] if len(row) > 5 else ""),
                hours_expected=_cell_float(row[6] if len(row) > 6 else 0),
                ects=_cell_float(row[2] if len(row) > 2 else 0),
            )
        )
    return out


_OFFICE_THEME_RGB = (
    "FFFFFF",
    "000000",
    "E7E6E6",
    "44546A",
    "4472C4",
    "ED7D31",
    "A5A5A5",
    "FFC000",
    "5B9BD5",
    "70AD47",
)


def _apply_tint(hex6: str, tint: float) -> str:
    r, g, b = (int(hex6[i : i + 2], 16) for i in (0, 2, 4))
    if tint < 0:
        r = int(r * (1.0 + tint))
        g = int(g * (1.0 + tint))
        b = int(b * (1.0 + tint))
    else:
        r = int(r + (255 - r) * tint)
        g = int(g + (255 - g) * tint)
        b = int(b + (255 - b) * tint)
    return (
        f"{max(0, min(255, r)):02X}"
        f"{max(0, min(255, g)):02X}"
        f"{max(0, min(255, b)):02X}"
    )


def _cell_fill_hex(cell: Any) -> str:
    try:
        fill = cell.fill
        if fill is None or fill.fill_type in (None, "none"):
            return ""
        fg = fill.fgColor
        if fg is None:
            return ""
        hex6 = ""
        if fg.type == "rgb" and fg.rgb:
            raw = str(fg.rgb).upper()
            if len(raw) == 8:
                hex6 = raw[2:]
            elif len(raw) == 6:
                hex6 = raw
        elif fg.type == "indexed" and fg.indexed is not None:
            from openpyxl.styles.colors import COLOR_INDEX

            idx = int(fg.indexed)
            if 0 <= idx < len(COLOR_INDEX):
                raw = str(COLOR_INDEX[idx]).upper()
                hex6 = raw[2:] if len(raw) == 8 else raw
        elif fg.type == "theme" and fg.theme is not None:
            idx = int(fg.theme)
            if 0 <= idx < len(_OFFICE_THEME_RGB):
                hex6 = _OFFICE_THEME_RGB[idx]
                tint = float(fg.tint or 0.0)
                if tint:
                    hex6 = _apply_tint(hex6, tint)
        if len(hex6) != 6:
            return ""
        if hex6 in ("FFFFFF", "000000"):
            return ""
        return hex6
    except Exception:
        return ""


def parse_grid_sheet(
    rows: list[tuple[Any, ...]],
    *,
    level: str,
    track: str,
    period: str,
    cell_fills: dict[tuple[int, int], str] | None = None,
) -> list[TimetableSlotRow]:
    week_cols: dict[int, tuple[str, int]] = {}
    week_row_idx = None
    for i, row in enumerate(rows):
        if len(row) > 1 and _cell_str(row[1]).startswith("Week"):
            week_row_idx = i
            for j, cell in enumerate(row):
                lab = _cell_str(cell)
                if lab.startswith("Week"):
                    num_m = re.search(r"(\d+)", lab)
                    week_cols[j] = (lab, int(num_m.group(1)) if num_m else 0)
            break
    if week_row_idx is None or not week_cols:
        return []

    out: list[TimetableSlotRow] = []
    current_day = ""
    day_dates: dict[int, str] = {}

    for row_offset, row in enumerate(rows[week_row_idx + 1 :]):
        row_idx = week_row_idx + 1 + row_offset
        c0 = _cell_str(row[0] if row else "")
        if c0 in _WEEKDAYS:
            current_day = c0
            day_dates = {}
            for j in week_cols:
                if j < len(row):
                    day_dates[j] = _cell_str(row[j])
            continue

        time_slot = normalize_time_slot(c0)
        if not time_slot or not current_day:
            continue

        for j, (week_label, week_number) in week_cols.items():
            raw = _cell_str(row[j] if j < len(row) else "")
            if not raw:
                continue
            legacy, mne, teacher = _parse_cell_content(raw)
            kind = _classify_cell(raw, legacy)
            fill_color = (cell_fills or {}).get((row_idx, j), "")
            out.append(
                TimetableSlotRow(
                    level=level,
                    track=track,
                    period=period,
                    week_label=week_label,
                    week_number=week_number,
                    week_start_date=day_dates.get(j, ""),
                    day_of_week=current_day,
                    time_slot=time_slot,
                    raw_text=raw,
                    legacy_code=legacy,
                    mne_module_code=mne,
                    teacher_initials=teacher,
                    room="",
                    slot_kind=kind,
                    fill_color=fill_color,
                )
            )
    return out


def parse_grid_sheet_from_worksheet(
    ws: Any,
    *,
    level: str,
    track: str,
    period: str,
) -> list[TimetableSlotRow]:
    all_rows = list(ws.iter_rows())
    value_rows: list[tuple[Any, ...]] = []
    cell_fills: dict[tuple[int, int], str] = {}
    for i, row in enumerate(all_rows):
        values: list[Any] = []
        for j, cell in enumerate(row):
            values.append(cell.value)
            fill = _cell_fill_hex(cell)
            if fill:
                cell_fills[(i, j)] = fill
        value_rows.append(tuple(values))
    return parse_grid_sheet(
        value_rows,
        level=level,
        track=track,
        period=period,
        cell_fills=cell_fills,
    )


def load_timetable_workbook(path: Path | str, *, academic_year: str = "") -> TimetableImportResult:
    from openpyxl import load_workbook

    p = Path(path)
    with warnings.catch_warnings():
        warnings.filterwarnings(
            "ignore",
            message="DrawingML support is incomplete.*",
            category=UserWarning,
            module="openpyxl.reader.drawings",
        )
        wb = load_workbook(p, read_only=False, data_only=False)
    try:
        sheet_names = list(wb.sheetnames)
        title_cell = ""
        if sheet_names:
            ws0 = wb[sheet_names[0]]
            first_rows = list(ws0.iter_rows(max_row=1, max_col=1, values_only=True))
            if first_rows and first_rows[0]:
                title_cell = _cell_str(first_rows[0][0])

        year = (academic_year or "").strip() or extract_academic_year_from_timetable(
            p, title_cell, sheet_names=sheet_names
        )
        level = "M1"
        for name in wb.sheetnames:
            meta = _parse_grid_sheet_meta(name) or _parse_m2_common_grid_sheet_meta(name)
            if meta:
                level = meta[0]
                break

        result = TimetableImportResult(
            academic_year=year,
            level=level,
            source_filename=p.name,
        )
        if not year:
            result.warnings.append(
                "Millésime non détecté — renseignez-le à l'import ou nommez le fichier « …2024-2025… »."
            )

        for name in wb.sheetnames:
            period = _parse_supervisor_sheet_meta(name)
            if period:
                ws = wb[name]
                rows = list(ws.iter_rows(values_only=True))
                result.reference_courses.extend(parse_supervisor_sheet(rows, period))
            elif _is_generic_supervisor_sheet(name):
                ws = wb[name]
                rows = list(ws.iter_rows(values_only=True))
                result.reference_courses.extend(parse_supervisor_sheet(rows))
            elif _is_m2_courses_sheet(name):
                ws = wb[name]
                rows = list(ws.iter_rows(values_only=True))
                result.reference_courses.extend(parse_m2_courses_sheet(rows))

        for name in wb.sheetnames:
            meta = _parse_grid_sheet_meta(name) or _parse_m2_common_grid_sheet_meta(name)
            if not meta:
                continue
            lv, track, period = meta
            ws = wb[name]
            result.slots.extend(
                parse_grid_sheet_from_worksheet(ws, level=lv, track=track, period=period)
            )

        if not result.slots:
            result.warnings.append("Aucune grille reconnue (onglets « M1 Physics S1 », etc.).")
        return result
    finally:
        wb.close()


def slot_hours(time_slot: str) -> float:
    return _SLOT_HOURS.get(normalize_time_slot(time_slot) or time_slot, 3.25)
