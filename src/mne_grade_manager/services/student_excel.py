"""Export / modèles Excel pour la liste d'étudiants (alignés sur la fiche étudiant)."""

from __future__ import annotations

import re
import unicodedata
from pathlib import Path
from typing import Any

from .dates import format_age_display
from .lookups import gender_label_fr

# Clés internes (base SQLite), dans l'ordre des colonnes Excel.
STUDENT_IMPORT_FIELD_KEYS: tuple[str, ...] = (
    "student_number_ine",
    "student_number_local",
    "last_name",
    "first_name",
    "gender",
    "birth_date",
    "nationality",
    "birth_place",
    "origin_institution",
    "origin_institution_country",
    "highest_diploma",
    "email_personal",
    "email_institutional",
    "phone",
    "enrollment_institution",
    "application_platform",
    "mon_master_ranking",
    "funding",
    "funding_other",
    "accommodations",
    "accommodations_other",
    "notes",
    "level",
    "track",
    "academic_year",
)

STUDENT_EXPORT_FIELD_KEYS: tuple[str, ...] = ("student_number",) + STUDENT_IMPORT_FIELD_KEYS

# Libellés identiques à la fiche étudiant (StudentProfileDialog).
STUDENT_FIELD_LABEL_FR: dict[str, str] = {
    "student_number": "Identifiant interne (base)",
    "student_number_ine": "N° I.N.E.",
    "student_number_local": "N° inscription (établissement)",
    "last_name": "Nom",
    "first_name": "Prénom",
    "gender": "Genre",
    "birth_date": "Date de naissance",
    "nationality": "Nationalité",
    "birth_place": "Lieu de naissance",
    "origin_institution": "Établissement d'origine",
    "origin_institution_country": "Pays (origine)",
    "highest_diploma": "Plus haut diplôme actuel",
    "email_personal": "Email personnel",
    "email_institutional": "Email institutionnel",
    "phone": "Téléphone",
    "enrollment_institution": "Établissement d'inscription",
    "application_platform": "Plateforme candidature",
    "mon_master_ranking": "Classement Mon Master",
    "funding": "Bourses / exemptions",
    "funding_other": "Autre bourse ou exemption",
    "accommodations": "Aménagements",
    "accommodations_other": "Autres aménagements",
    "notes": "Notes",
    "level": "Niveau",
    "track": "Parcours",
    "academic_year": "Année universitaire",
}

STUDENT_REQUIRED_IMPORT_KEYS: tuple[str, ...] = (
    "student_number_ine",
    "student_number_local",
    "last_name",
    "first_name",
)

# Rétrocompatibilité (anciens modèles avec noms techniques ou variantes).
_HEADER_ALIASES_TO_KEY: dict[str, str] = {
    "student_number": "student_number",
    "student_number_ine": "student_number_ine",
    "ine": "student_number_ine",
    "numero_ine": "student_number_ine",
    "n_ine": "student_number_ine",
    "n_i_n_e": "student_number_ine",
    "i_n_e": "student_number_ine",
    "student_number_local": "student_number_local",
    "local_student_number": "student_number_local",
    "numero_etablissement": "student_number_local",
    "numero_etudiant_etablissement": "student_number_local",
    "n_etablissement": "student_number_local",
    "n_etablissement_apogee": "student_number_local",
    "apogee": "student_number_local",
    "n_apogee": "student_number_local",
    "last_name": "last_name",
    "lastname": "last_name",
    "nom": "last_name",
    "surname": "last_name",
    "first_name": "first_name",
    "firstname": "first_name",
    "prenom": "first_name",
    "given_name": "first_name",
    "gender": "gender",
    "genre": "gender",
    "sexe": "gender",
    "birth_date": "birth_date",
    "date_naissance": "birth_date",
    "naissance": "birth_date",
    "nationality": "nationality",
    "nationalite": "nationality",
    "birth_place": "birth_place",
    "lieu_naissance": "birth_place",
    "origin_institution": "origin_institution",
    "etablissement_origine": "origin_institution",
    "origin_institution_country": "origin_institution_country",
    "pays_etablissement_origine": "origin_institution_country",
    "pays_origine": "origin_institution_country",
    "highest_diploma": "highest_diploma",
    "plus_haut_diplome": "highest_diploma",
    "plus_haut_diplome_actuel": "highest_diploma",
    "diplome_actuel": "highest_diploma",
    "email_personal": "email_personal",
    "email_perso": "email_personal",
    "email": "email_personal",
    "mail": "email_personal",
    "email_institutional": "email_institutional",
    "email_inst": "email_institutional",
    "email_institutionnel": "email_institutional",
    "phone": "phone",
    "telephone": "phone",
    "tel": "phone",
    "mobile": "phone",
    "telephone_portable": "phone",
    "enrollment_institution": "enrollment_institution",
    "etablissement_inscription": "enrollment_institution",
    "etablissement": "enrollment_institution",
    "application_platform": "application_platform",
    "plateforme_candidature": "application_platform",
    "mon_master_ranking": "mon_master_ranking",
    "classement_mon_master": "mon_master_ranking",
    "classement_monmaster": "mon_master_ranking",
    "ranking_mon_master": "mon_master_ranking",
    "funding": "funding",
    "bourses": "funding",
    "bourse": "funding",
    "bourses_exemptions": "funding",
    "exemption_frais": "funding",
    "exemption_frais_inscription": "funding",
    "funding_other": "funding_other",
    "autre_bourse": "funding_other",
    "autres_bourses": "funding_other",
    "accommodations": "accommodations",
    "amenagement": "accommodations",
    "amenagements": "accommodations",
    "accommodations_other": "accommodations_other",
    "amenagement_autres": "accommodations_other",
    "autres_amenagements": "accommodations_other",
    "notes": "notes",
    "commentaire": "notes",
    "commentaires": "notes",
    "level": "level",
    "annee": "level",
    "year": "level",
    "m1m2": "level",
    "niveau": "level",
    "track": "track",
    "parcours": "track",
    "specialite": "track",
    "academic_year": "academic_year",
    "annee_universitaire": "academic_year",
    "m1_c": "m1_c",
    "m1c": "m1_c",
    "m1_p": "m1_p",
    "m1p": "m1p",
}

# Anciens noms d'en-têtes (compatibilité code existant).
STUDENT_IMPORT_HEADERS = STUDENT_IMPORT_FIELD_KEYS
STUDENT_EXPORT_HEADERS = STUDENT_EXPORT_FIELD_KEYS
STUDENT_EXCEL_HEADER_LABELS_FR = {
    key: f"{STUDENT_FIELD_LABEL_FR[key]} — {hint}"
    for key, hint in {
        "student_number": "export uniquement, ex. MNE-DUPONT-JE-A7K2",
        "student_number_ine": "optionnel",
        "student_number_local": "optionnel — attribué par l'établissement à l'inscription",
        "last_name": "obligatoire à l'import",
        "first_name": "obligatoire à l'import",
        "gender": "M, F, O ou Homme / Femme / Autre",
        "birth_date": "AAAA-MM-JJ ou JJ/MM/AAAA",
        "nationality": "",
        "birth_place": "",
        "origin_institution": "",
        "origin_institution_country": "",
        "highest_diploma": "ex. Licence Physique (Bac+3)",
        "email_personal": "",
        "email_institutional": "@universite-paris-saclay.fr, @ip-paris.fr, etc.",
        "phone": "numéro de téléphone (portable ou fixe)",
        "enrollment_institution": "",
        "application_platform": "MonMaster, UPSay, IPParis…",
        "mon_master_ranking": "ex. 1, 10, NC — optionnel",
        "funding": "campus_france, idex, eiffel, tuition_exemption (virgules)",
        "funding_other": "texte libre",
        "accommodations": "tiers_temps, salle_isolee, pc (virgules)",
        "accommodations_other": "texte libre",
        "notes": "commentaires",
        "level": "M1 ou M2",
        "track": "P, C, NPD, NPO, DWM, NFC, NRPE",
        "academic_year": "ex. 2025-2026",
    }.items()
}

STUDENT_EXCEL_EXAMPLE_ROW: tuple[str, ...] = (
    "18012345678",
    "EN00005920",
    "Dupont",
    "Marie",
    "F",
    "2000-05-15",
    "France",
    "Lyon",
    "Université Claude Bernard Lyon 1",
    "France",
    "Licence Physique (Bac+3)",
    "marie.dupont@gmail.com",
    "marie.dupont@universite-paris-saclay.fr",
    "06 12 34 56 78",
    "Université Paris-Saclay",
    "MonMaster",
    "",
    "campus_france",
    "",
    "tiers_temps",
    "",
    "",
    "M1",
    "P",
    "2025-2026",
)

_INSTRUCTIONS_ROWS: tuple[tuple[str, str], ...] = (
    (
        "Colonnes obligatoires",
        "Nom, Prénom",
    ),
    (
        "Identifiant interne (base)",
        "Non présent dans le modèle d'import : généré par l'application, visible à l'export.",
    ),
    ("Email institutionnel", "Domaine accepté : Paris-Saclay, ENSTA, IP Paris, Chimie Paris."),
    ("Photo / PDF", "Non importables via Excel : fiche étudiant → Photo & documents."),
    ("Import", "Onglet Étudiants → Importer Excel (.xlsx)."),
    ("Export", "Mêmes en-têtes que la fiche étudiant, pour réimport ou mise à jour."),
)


def normalize_mon_master_ranking(raw: Any) -> str:
    """Texte libre (rang, « NC », etc.) ; les nombres Excel sont normalisés."""
    if raw is None:
        return ""
    if isinstance(raw, bool):
        return ""
    if isinstance(raw, int):
        return str(raw)
    if isinstance(raw, float):
        if raw == int(raw):
            return str(int(raw))
        return str(raw).replace(".", ",")
    s = str(raw).strip()
    if not s:
        return ""
    if s.upper() == "NC":
        return "NC"
    try:
        f = float(s.replace(",", "."))
        if f == int(f):
            return str(int(f))
    except ValueError:
        pass
    return s


def field_label_fr(key: str) -> str:
    return STUDENT_FIELD_LABEL_FR.get(key, key)


def excel_column_headers(*, for_export: bool) -> list[str]:
    keys = STUDENT_EXPORT_FIELD_KEYS if for_export else STUDENT_IMPORT_FIELD_KEYS
    return [field_label_fr(k) for k in keys]


def normalize_excel_header(value: Any) -> str:
    if value is None:
        return ""
    s = str(value).strip().lower()
    s = unicodedata.normalize("NFKD", s)
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    s = re.sub(r"[^a-z0-9]+", "_", s).strip("_")
    return s


def _build_header_lookup() -> dict[str, str]:
    lookup: dict[str, str] = dict(_HEADER_ALIASES_TO_KEY)
    for key, label in STUDENT_FIELD_LABEL_FR.items():
        lookup[normalize_excel_header(label)] = key
    return lookup


_HEADER_LOOKUP = _build_header_lookup()


def resolve_field_key_from_header(header_cell: Any) -> str | None:
    norm = normalize_excel_header(header_cell)
    if not norm:
        return None
    return _HEADER_LOOKUP.get(norm)


def build_import_column_map(header_row: tuple[Any, ...]) -> dict[str, int | None]:
    """Associe chaque clé interne à l'index de colonne (ou None si absente)."""
    indices: dict[str, int] = {}
    for idx, cell in enumerate(header_row):
        key = resolve_field_key_from_header(cell)
        if key and key not in indices:
            indices[key] = idx
    extra_keys = ("m1_c", "m1p")
    all_keys = set(STUDENT_EXPORT_FIELD_KEYS) | set(extra_keys)
    return {key: indices.get(key) for key in all_keys}


def student_dict_to_excel_row(s: dict[str, Any]) -> list[Any]:
    """Une ligne de données pour export (même ordre que les en-têtes français)."""
    g = str(s.get("gender") or "").strip()
    if g in {"M", "F", "O"}:
        gender_cell = g
    else:
        gender_cell = gender_label_fr(g) if g else ""
    return [
        s.get("student_number", ""),
        s.get("student_number_ine", ""),
        s.get("student_number_local", ""),
        s.get("last_name", ""),
        s.get("first_name", ""),
        gender_cell,
        str(s.get("birth_date") or "").strip(),
        s.get("nationality", ""),
        s.get("birth_place", ""),
        s.get("origin_institution", ""),
        s.get("origin_institution_country", ""),
        s.get("highest_diploma", ""),
        s.get("email_personal", ""),
        s.get("email_institutional", ""),
        s.get("phone", ""),
        s.get("enrollment_institution", ""),
        s.get("application_platform", ""),
        s.get("mon_master_ranking", ""),
        s.get("funding", ""),
        s.get("funding_other", ""),
        s.get("accommodations", ""),
        s.get("accommodations_other", ""),
        s.get("notes", ""),
        s.get("level", ""),
        s.get("track", ""),
        s.get("academic_year", ""),
    ]


def write_students_workbook(
    path: str | Path,
    students: list[dict[str, Any]],
    *,
    include_instructions: bool = False,
    include_example: bool = False,
) -> None:
    """Écrit un fichier Excel (export données ou modèle d'import)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    ws = wb.active
    ws.title = "Étudiants"
    for_export = bool(students)
    headers = excel_column_headers(for_export=for_export)
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    if include_example:
        ws.append(list(STUDENT_EXCEL_EXAMPLE_ROW))
    for s in students:
        ws.append(student_dict_to_excel_row(s))
    ws.freeze_panes = "A2"

    if include_instructions:
        ins = wb.create_sheet("Instructions")
        ins.append(["Colonne (fiche étudiant)", "Description"])
        ins["A1"].font = Font(bold=True)
        ins["B1"].font = Font(bold=True)
        field_keys = STUDENT_EXPORT_FIELD_KEYS if for_export else STUDENT_IMPORT_FIELD_KEYS
        for key in field_keys:
            desc = STUDENT_EXCEL_HEADER_LABELS_FR.get(key, "")
            ins.append([field_label_fr(key), desc.split(" — ", 1)[-1] if desc else ""])
        ins.append([])
        ins.append(["Remarques", ""])
        for title, text in _INSTRUCTIONS_ROWS:
            ins.append([title, text])

    wb.save(str(path))


def write_student_import_template(path: str | Path) -> None:
    """Modèle vide prêt à remplir + ligne d'exemple + onglet Instructions."""
    write_students_workbook(
        path,
        [],
        include_instructions=True,
        include_example=True,
    )
